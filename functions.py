import bisect
import copy

from ct import *
from ov import *
from item import *
from of import *
from cliente import *
import pandas as pd
import time
import math
import numpy as np
import datetime
from fractions import Fraction
from tqdm import tqdm
import time
import shutil
import codecs
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import os
import json
import warnings
warnings.filterwarnings("ignore")
import copy
method=0
from datetime import date

def verificar_referencias_c():

    df_cs = pd.read_csv('data/20. referencias c.csv', encoding='ISO-8859-1', sep=",",error_bad_lines=False)
    df_cs['material'] = df_cs['material'].astype(str)
    cs = df_cs['material'].tolist()

    # IMPORTAR FATORES DE CONVERSAO
    df_fi = pd.read_csv('data/23. fatores_conversao.csv', sep=",", encoding="ISO-8859-1",error_bad_lines=False)
    materiais_fi = df_fi['Material'].tolist()
    componentes_fi = df_fi['Componente'].tolist()
    fi = df_fi['FI'].tolist()

    for of in ofs:
        material_of=of.codigo_material
        for index in range(0,len(materiais_fi)):
            if material_of==materiais_fi[index]:
                if componentes_fi[index] in cs:
                    of.is_c=1


def find_index(lst, a):

    return [i for i, x in enumerate(lst) if x==a]


def ler_ocupacao_aglomerados():

    #VER REFERENCIAS C
    df_cs = pd.read_csv('data/20. referencias c.csv', encoding='ISO-8859-1', sep=",",error_bad_lines=False)
    df_cs['material'] = df_cs['material'].astype(str)
    cs = df_cs['material'].tolist()
    quantidade_cs= [[0]*53]*len(cs)

    #VER OFS PLANEADAS DE EMBALAGEM
    df_ofs=pd.read_csv('data/08.  ofs.csv',sep=",",encoding='ISO-8859-1',error_bad_lines=False)
    df_ofs=df_ofs[(df_ofs['Ordem de produção / planeada']>1600000000)&(df_ofs['Ordem de produção / planeada']<1700000000)]
    df_ofs=df_ofs[df_ofs['Descritivo de Centro de Trabalho'].str.contains('EMB')]
    df_ofs['Data-base do fim'] = pd.to_datetime(df_ofs['Data-base do fim'], format='%d/%m/%Y',
                                                          errors='coerce')
    last_week = date(datetime.datetime.today().year , 12, 28)
    n_weeks=last_week.isocalendar()[1]-semana_inicio_plano

    df_ofs['semana_atual']=df_ofs['Data-base do fim'].apply(lambda x: x.isocalendar()[1])
    #df_ofs['semana'] = df_ofs['Data-base do fim'].apply(lambda x: x.isocalendar()[1]-semana_inicio_plano if x.isocalendar()[0]==datetime.datetime.today().year else n_weeks+x.isocalendar()[1])
    df_ofs['semana'] = df_ofs['Data-base do fim'].apply(lambda x: x.isocalendar()[1]-semana_inicio_plano )
    df_ofs=df_ofs.sort_values(by=['semana_atual'],ascending=True)
    df_ofs=df_ofs[df_ofs['semana']>=0]
    ofs=df_ofs['Ordem de produção / planeada'].tolist()
    descricao=df_ofs['Texto breve de material'].tolist()
    materiais=df_ofs['Material de produção'].tolist()
    semana=df_ofs['semana'].tolist() #VERIFICAR
    quantidade=df_ofs['Quantidade total da ordem'].tolist()

    #IMPORTAR FATORES DE CONVERSAO
    df_fi=pd.read_csv('data/23. fatores_conversao.csv',sep=",",encoding="ISO-8859-1",error_bad_lines=False)
    materiais_fi=df_fi['Material'].tolist()
    componentes_fi=df_fi['Componente'].tolist()
    fi=df_fi['FI'].tolist()

    #IMPORTAR CADENCIAS
    df_bom=pd.read_csv('data/19. BOM.csv',sep=",",encoding="ISO-8859-1",error_bad_lines=False)
    df_bom=df_bom[df_bom['VP']==1]
    df_bom=df_bom[['Material','Cadência máquina','Centro trabalho','Descrição Operação']]
    df_bom=df_bom.dropna()
    df_bom=df_bom.drop_duplicates()
    materiais_bom=df_bom['Material'].tolist()
    cadencias_bom=df_bom['Cadência máquina'].tolist()
    ct_bom=df_bom['Centro trabalho'].tolist()
    descritivo_operacao_bom=df_bom['Descrição Operação'].tolist()

    test=[]

    for index1 in range(len(materiais)):
        material=materiais[index1]
        for index2 in range(len(materiais_fi)):
            material_fi=materiais_fi[index2]
            if material==material_fi:
                #verificar centro de trabalho da precedencia
                material_componente_atual=componentes_fi[index2]
                centro_de_trabalho=""
                for index3 in range(len(materiais_bom)):
                    material_componente=materiais_bom[index3]
                    if material_componente==material_componente_atual:
                        if 'AGL' in descritivo_operacao_bom[index3]:
                            centro_de_trabalho=ct_bom[index3]
                            cadencia=cadencias_bom[index3]
                        break
                for index4 in range(len(cs)):
                    if material_componente_atual==cs[index4]:
                        temp = quantidade_cs[index4].copy()
                        temp[semana[index1]]+=quantidade[index1]*float(fi[index2])
                        quantidade_cs[index4]=temp.copy()
                if centro_de_trabalho!="":
                    for ct in cts:
                        if ct.nome==centro_de_trabalho:
                            duracao_agl=quantidade[index1]*float(fi[index2])/cadencia*60
                            new_row={'of':ofs[index1],'descricao':descricao[index1],'semana':semana[index1]+semana_inicio_plano,'ct':ct.nome,'quantidade agl':duracao_agl}
                            test.append(new_row)
                            ct.capacidade[semana[index1]]=ct.capacidade[semana[index1]]-duracao_agl

    df_teste=pd.DataFrame(test)
    df_teste.to_csv('data/test.csv',encoding='ISO-8859-1')
    output=[]
    for index in range(len(quantidade_cs)):
        c=quantidade_cs[index]
        for semana in range(len(c)):
            qtd=c[semana]
            new_row={'c':cs[index],'semana':semana+semana_inicio_plano,'quantidade':qtd}
            output.append(new_row)

    df=pd.DataFrame(output)
    df.to_csv('data/22. quantidade cs.csv',sep=",",encoding='ISO-8859-1')

def get_next_componente(precedencia,materiais_unique,precendencias,df_bom,fatores_incorporacao):

    result_to_return=[]

    try:
        id=-1
        for posicao in range(len(materiais_unique)):
            if float(precedencia)==materiais_unique[posicao]:
                id=posicao
                break

        ct=df_bom[df_bom['Material']==(precedencia)]['Centro trabalho'].tolist()
        descricao = df_bom[df_bom['Material'] == (precedencia)]['Descrição Operação'].tolist()

        if len(ct) != 0:
            if 'AG' in ct[0] and 'AGL' in descricao[0]:
                result_to_return.append(1)
                result_to_return.append(precedencia)
            else:
                for precendencia_to_append in precendencias[id]:
                    result_to_return.append(precendencia_to_append)
        else:
            result_to_return.append(-1)
            result_to_return.append(precedencia)
    except:
        result_to_return.append(1)
        result_to_return.append(precedencia)


    return result_to_return,id

def verificar_agl(id_material,precendencias,df_bom,materiais_unique,resultado,fatores_incorporacao):

    resultados_to_return=[]
    if len(resultado)==0:
        resultado=precendencias[id_material]
    condicao=False

    while condicao==False:
        if len(resultado)==1:
            resultado,id_material=get_next_componente(precendencias[id_material][0],materiais_unique,precendencias,df_bom,fatores_incorporacao)
        else:
            for precedencia in resultado:
                id_precedencia = -1
                for id_posicao in range(len(materiais_unique)):
                    if materiais_unique[id_posicao]==(precedencia):
                        id_precedencia=id_posicao
                        break
                if id_precedencia!=-1:
                    resultado_precedencia_atual,condicao=verificar_agl(id_precedencia,precendencias,df_bom,materiais_unique,precendencias[id_precedencia],fatores_incorporacao)

        if len(resultado)>0:
            if resultado[0]==1 or resultado[0]==-1:
                condicao=True
            if resultado[0]==1:
                resultados_to_return.append(resultado[1])

    return resultados_to_return,condicao

def ler_cts_embalagem():

    global lista_materials

    df=pd.read_csv('data/17. cts embalagem.csv',encoding='iso-8859-1',error_bad_lines=False)

    lista_materials=df['Material'].tolist()

    return lista_materials

def atualizar_ficheiros():
    df1 = pd.read_excel('ACC_Ordens por planear - Kaizen.xlsx',error_bad_lines=False)
    df2 = pd.read_csv('data/14.  ofs after plan.csv',error_bad_lines=False)
    if df1.equals(df2):
        return -1
    elif method == 0:

        read_file = pd.read_excel("ACC_Ordens por planear - Kaizen.xlsx",error_bad_lines=False)
        read_file.to_csv("data/14.  ofs after plan.csv", index=None, header=True, date_format='%d/%m/%Y',encoding='utf-8-sig',error_bad_lines=False)

    else:

        read_file = pd.read_excel("ACC_Ordens por planear - Kaizen.xlsx",error_bad_lines=False)
        read_file.to_csv("data/14.  ofs after plan.csv", index=None, header=True, date_format='%d/%m/%Y',error_bad_lines=False)
        read_file.to_csv("data/08.  ofs.csv", index=None, header=True, date_format='%d/%m/%Y',encoding='utf-8-sig',error_bad_lines=False)

    return 0

def find(lst, a):
    result = []
    for i, x in enumerate(lst):
        if x==a:
            result.append(i)
    return result

print('importar clientes')
start=time.time()

def importar_clientes():

    clientes=[]
    df_clientes=pd.read_csv('data/01. clientes.csv',sep=",",encoding='iso-8859-1',error_bad_lines=False)
    df_clientes['Carga Completa']=df_clientes['Carga Completa'].fillna(1)
    df_clientes['Lead Time Acordado (semanas)'] = df_clientes['Lead Time Acordado (semanas)'].fillna(0)
    df_clientes_total=df_clientes.copy()
    df_clientes=df_clientes.drop_duplicates(['Grupo'])
    nome_clientes=df_clientes['Nome'].tolist()
    carga_completa=df_clientes['Carga Completa'].tolist()
    lead_time=df_clientes['Lead Time Acordado (semanas)'].tolist()
    grupo=df_clientes['Grupo'].tolist()

    for index in range(len(nome_clientes)):
        new_cliente=cliente(index,grupo[index],lead_time[index],carga_completa[index])
        clientes.append(new_cliente)

    with tqdm(total=len(clientes)) as pbar:

        for index in range(len(clientes)):

            grupo=clientes[index].nome

            df_grupo=df_clientes_total[df_clientes_total['Grupo']==grupo]

            nomes_total = df_grupo['Nome'].tolist()

            for pos_nome in range(len(nomes_total)):

                clientes[index].descricao_clientes.append(nomes_total[pos_nome])

            pbar.update(1)

    pbar.close()

    return clientes

clientes=importar_clientes()

print('importar centros de trabalho')

def export_capacidades_in():

    rows=[]

    global cts
    global semana_inicio_plano

    ler_ocupacao_aglomerados()

    for index in range(len(cts)):

        vetor_capacidades=cts[index].capacidade_iniciais

        for pos_semana in range(0,len(vetor_capacidades)):

            new_row={'area':cts[index].area,'centro de trabalho':cts[index].nome,'acabamento':cts[index].acabamento,'semana':pos_semana+semana_inicio_plano,'capacidade':cts[index].capacidade_iniciais[pos_semana],'ocupacao inicial':cts[index].capacidade[pos_semana]}
            rows.append(new_row)

    df_rows=pd.DataFrame(rows)
    df_rows.to_csv('data/101. capacidade inicial.csv', encoding='utf-8')

def importar_cts(caminho,df_blocos,df_virados,df_clientes,cts,area,method):

    try:

        #todo adicionar caminho

        df_cnm=pd.read_excel('G:/Operacoes/Logistica/PLANEAMENTO/1. PlaneamentoTOTAL/'+str(caminho),sheet_name='Mapa SAP')

    except:

        df_cnm = pd.read_excel(caminho, sheet_name='Mapa SAP')

    df_acabamentos = pd.read_csv('data/07.  acabamentos.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
    tipos_de_acabamento=df_acabamentos['tipo'].tolist()
    acabamentos=set(tipos_de_acabamento)
    acabamentos = (list(acabamentos))


    dias_semana=df_cnm.iloc[6:, 2]
    ocupacoes=df_cnm.iloc[6:,3:].values.tolist()
    dias_semana.index = np.arange(0, len(dias_semana))
    centros_trabalho=df_cnm.iloc[1:6,3:].T.dropna()
    excecoes=df_cnm.iloc[61:,:]
    excecoes.columns = excecoes.iloc[0]
    excecoes = excecoes[1:]
    excecoes.dropna(axis=1)
    ct_excecao=excecoes['CT'].tolist()
    cc_excecao=excecoes['CC'].tolist()
    de_excecao=excecoes['DE'].tolist()
    ate_excecao=excecoes['ATÉ'].tolist()
    n_turnos_excecao=excecoes['N.º TURNOS'].tolist()

    areas=centros_trabalho.iloc[:,0].tolist()
    turnos=centros_trabalho.iloc[:,1].tolist()
    horas_por_turno=centros_trabalho.iloc[:,2].tolist()
    oee=centros_trabalho.iloc[:,3].tolist()
    cts_nomes=centros_trabalho.iloc[:,4].tolist()

    semana_inicio_plano = datetime.datetime.now().isocalendar()[1]
    if method == 0:
        semana_inicio_plano = semana_inicio_plano + 1
    else:
        semana_inicio_plano = semana_inicio_plano + 4

    ct_blocos = df_blocos['ct'].tolist()
    capacidade_blocos = df_blocos['blocos semana'].tolist()

    ct_virados = df_virados['ct'].tolist()
    capacidade_virados = df_virados['blocos semana'].tolist()

    ct_clientes = df_clientes['ct'].tolist()
    grupo_cliente = df_clientes['cliente'].tolist()
    capacidade_cliente = df_clientes['capacidade'].tolist()

    counter=len(cts)
    helper=[]

    with tqdm(total=len(cts_nomes)) as pbar:

        for index in range(len(cts_nomes)):

            capacidade_bl = 9999999999
            capacidade_vir = 9999999999

            if (str(cts_nomes[index]) not in helper or ((cts_nomes[index] in helper and areas[index] in acabamentos)
            and cts_nomes[index]) != 'CCSBALJT' and cts_nomes[index] != 'CCSEMBME') and cts_nomes[index][:3]==area and areas[index]!='Total' and cts_nomes[index]!="CRMEMBPL":

                if areas[index] in acabamentos:

                    new_ct = ct(counter,cts_nomes[index],areas[index],area)

                else:

                    new_ct = ct(counter, cts_nomes[index], "Total", area)

                cts.append(new_ct)

                if cts_nomes[index] in ct_blocos:
                    posicao_blocos = ct_blocos.index(cts_nomes[index])
                    capacidade_bl = capacidade_blocos[posicao_blocos]

                if cts_nomes[index] in ct_virados:
                    posicao_virados = ct_virados.index(cts_nomes[index])
                    capacidade_vir = capacidade_virados[posicao_virados]

                for index_semana in range(semana_inicio_plano,53):

                    n_turnos = turnos[index]

                    for index_excecao in range(len(ct_excecao)):
                        if ct_excecao[index_excecao] == cts_nomes[index] and cc_excecao[index_excecao] == areas[
                            index] and index_semana >= de_excecao[index_excecao] and index_semana <= ate_excecao[
                            index_excecao]:
                            n_turnos = n_turnos_excecao[index_excecao]
                            break

                    try:
                        capacidade = int(dias_semana[index_semana-1]) * n_turnos * oee[index] * 60 * horas_por_turno[
                            index]
                    except:
                        capacidade = 0

                    new_ct.capacidade.append(capacidade)
                    new_ct.capacidade_virados.append(capacidade_vir)
                    new_ct.capacidade_blocos.append(capacidade_bl)
                    new_ct.capacidade_iniciais.append(capacidade)

                    try:

                        new_ct.capacidade[-1]=new_ct.capacidade[-1]-ocupacoes[index_semana-1][index]*capacidade

                    except:
                        new_ct.capacidade[-1] = new_ct.capacidade[-1]


                for posicao in range(len(clientes)):
                    capacidade_cliente_atual = [0]*len(new_ct.capacidade_iniciais)
                    new_ct.capacidade_clientes.append(capacidade_cliente_atual)
                    new_ct.capacidade_iniciais_clientes.append(capacidade_cliente_atual)

                if cts_nomes[index] in ct_clientes:

                    posicao_clientes=find(ct_clientes,cts_nomes[index])

                    for id_posicao in range(len(posicao_clientes)):

                        posicao=posicao_clientes[id_posicao]

                        grupo_tofind=grupo_cliente[posicao]
                        capacidade=capacidade_cliente[posicao]

                        for id_cliente in range(len(clientes)):

                            if clientes[id_cliente].nome==grupo_tofind:

                                capacidade_cliente_atual = [i * capacidade for i in new_ct.capacidade_iniciais]

                                new_ct.capacidade_clientes[id_cliente]=capacidade_cliente_atual
                                new_ct.capacidade_iniciais_clientes[id_cliente] = capacidade_cliente_atual

                counter+=1

                helper.append(str(cts_nomes[index]))

            pbar.update(1)

        pbar.close()

    return cts

def importar_cts_ccs(caminho,df_blocos,df_virados,df_clientes,cts,area,method):

    try:

        #todo adicionar caminho

        df_cnm=pd.read_excel('G:/Operacoes/Logistica/PLANEAMENTO/1. PlaneamentoTOTAL/'+str(caminho),sheet_name='Mapa SAP')

    except:

        df_cnm = pd.read_excel(caminho, sheet_name='Mapa SAP')

    dias_semana=df_cnm.iloc[6:, 2]
    ocupacoes=df_cnm.iloc[6:,3:].values.tolist()
    dias_semana.index = np.arange(0, len(dias_semana))
    centros_trabalho=df_cnm.iloc[1:6,3:].T.dropna()
    excecoes=df_cnm.iloc[61:,:]
    excecoes.columns = excecoes.iloc[0]
    excecoes = excecoes[1:]
    excecoes.dropna(axis=1)
    ct_excecao=excecoes['CT'].tolist()
    cc_excecao=excecoes['CC'].tolist()
    de_excecao=excecoes['DE'].tolist()
    ate_excecao=excecoes['ATÉ'].tolist()
    n_turnos_excecao=excecoes['N.º TURNOS'].tolist()

    areas=centros_trabalho.iloc[:,0].tolist()
    turnos=centros_trabalho.iloc[:,1].tolist()
    horas_por_turno=centros_trabalho.iloc[:,2].tolist()
    oee=centros_trabalho.iloc[:,3].tolist()
    cts_nomes=centros_trabalho.iloc[:,4].tolist()

    semana_inicio_plano = datetime.datetime.now().isocalendar()[1]
    if method == 0:
        semana_inicio_plano = semana_inicio_plano + 1
    else:
        semana_inicio_plano = semana_inicio_plano + 4

    ct_blocos = df_blocos['ct'].tolist()
    capacidade_blocos = df_blocos['blocos semana'].tolist()

    ct_virados = df_virados['ct'].tolist()
    capacidade_virados = df_virados['blocos semana'].tolist()

    ct_clientes = df_clientes['ct'].tolist()
    grupo_cliente = df_clientes['cliente'].tolist()
    capacidade_cliente = df_clientes['capacidade'].tolist()

    counter=len(cts)
    helper=[]

    with tqdm(total=len(cts_nomes)) as pbar:

        for index in range(len(cts_nomes)):

            capacidade_bl = 9999999999
            capacidade_vir = 9999999999

            if str(cts_nomes[index]) and cts_nomes[index][:3]==area:

                if cts_nomes[index]=="CCSBALJT" or cts_nomes[index]=="CCSEMBME":

                    new_ct = ct(counter,cts_nomes[index],areas[index],area)

                else:

                    new_ct = ct(counter, cts_nomes[index], "Total", area)

                cts.append(new_ct)

                if cts_nomes[index] in ct_blocos:
                    posicao_blocos = ct_blocos.index(cts_nomes[index])
                    capacidade_bl = capacidade_blocos[posicao_blocos]

                if cts_nomes[index] in ct_virados:
                    posicao_virados = ct_virados.index(cts_nomes[index])
                    capacidade_vir = capacidade_virados[posicao_virados]

                for index_semana in range(semana_inicio_plano,53):

                    n_turnos = turnos[index]

                    for index_excecao in range(len(ct_excecao)):
                        if ct_excecao[index_excecao] == cts_nomes[index] and cc_excecao[index_excecao] == areas[
                            index] and index_semana >= de_excecao[index_excecao] and index_semana <= ate_excecao[
                            index_excecao]:
                            n_turnos = n_turnos_excecao[index_excecao]
                            break

                    try:
                        capacidade = int(dias_semana[index_semana-1]) * n_turnos * oee[index] * 60 * horas_por_turno[
                            index]
                    except:
                        capacidade = 0

                    new_ct.capacidade.append(capacidade)
                    new_ct.capacidade_virados.append(capacidade_vir)
                    new_ct.capacidade_blocos.append(capacidade_bl)
                    new_ct.capacidade_iniciais.append(capacidade)

                    try:

                        new_ct.capacidade[-1]=new_ct.capacidade[-1]-ocupacoes[index_semana-1][index]*capacidade

                    except:
                        new_ct.capacidade[-1] = new_ct.capacidade[-1]


                for posicao in range(len(clientes)):
                    capacidade_cliente_atual = [0]*len(new_ct.capacidade_iniciais)
                    new_ct.capacidade_clientes.append(capacidade_cliente_atual)
                    new_ct.capacidade_iniciais_clientes.append(capacidade_cliente_atual)

                if cts_nomes[index] in ct_clientes:

                    posicao_clientes=find(ct_clientes,cts_nomes[index])

                    for id_posicao in range(len(posicao_clientes)):

                        posicao=posicao_clientes[id_posicao]

                        grupo_tofind=grupo_cliente[posicao]
                        capacidade=capacidade_cliente[posicao]

                        for id_cliente in range(len(clientes)):

                            if clientes[id_cliente].nome==grupo_tofind:

                                capacidade_cliente_atual = [i * capacidade for i in new_ct.capacidade_iniciais]

                                new_ct.capacidade_clientes[id_cliente]=capacidade_cliente_atual
                                new_ct.capacidade_iniciais_clientes[id_cliente] = capacidade_cliente_atual

                counter+=1

                helper.append(str(cts_nomes[index])+str(areas[index]))

            pbar.update(1)

        pbar.close()

    return cts

def adicionar_capacidade_cts(caminho,cts,area,method):

    try:

        #todo adicionar caminho

        df_cnm=pd.read_excel('G:/Operacoes/Logistica/PLANEAMENTO/1. PlaneamentoTOTAL/'+str(caminho),sheet_name='Mapa SAP')

    except:

        df_cnm = pd.read_excel(caminho, sheet_name='Mapa SAP')

    df_acabamentos = pd.read_csv('data/07.  acabamentos.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
    tipos_de_acabamento=df_acabamentos['tipo'].tolist()
    acabamentos=set(tipos_de_acabamento)
    acabamentos = (list(acabamentos))


    dias_semana=df_cnm.iloc[6:, 2]
    ocupacoes=df_cnm.iloc[6:,3:].values.tolist()
    dias_semana.index = np.arange(0, len(dias_semana))
    centros_trabalho=df_cnm.iloc[1:6,3:].T.dropna()
    excecoes=df_cnm.iloc[61:,:]
    excecoes.columns = excecoes.iloc[0]
    excecoes = excecoes[1:]
    excecoes.dropna(axis=1)
    ct_excecao=excecoes['CT'].tolist()
    cc_excecao=excecoes['CC'].tolist()
    de_excecao=excecoes['DE'].tolist()
    ate_excecao=excecoes['ATÉ'].tolist()
    n_turnos_excecao=excecoes['N.º TURNOS'].tolist()

    areas=centros_trabalho.iloc[:,0].tolist()
    turnos=centros_trabalho.iloc[:,1].tolist()
    horas_por_turno=centros_trabalho.iloc[:,2].tolist()
    oee=centros_trabalho.iloc[:,3].tolist()
    cts_nomes=centros_trabalho.iloc[:,4].tolist()

    ct_clientes = df_clientes['ct'].tolist()
    grupo_cliente = df_clientes['cliente'].tolist()
    capacidade_cliente = df_clientes['capacidade'].tolist()
    ids_used=[]
    counter=0
    helper=[]

    with tqdm(total=len(cts_nomes)) as pbar:

        for index in range(len(cts_nomes)):

            capacidade_bl = 9999999999
            capacidade_vir = 9999999999

            id_ct=-1
            for ct in cts:
                if ct.nome==cts_nomes[index] and areas[index]==ct.acabamento and (ct.id not in ids_used):
                    id_ct=ct.id
                    break
                elif ct.acabamento=="Total" and areas[index] not in acabamentos and ct.nome==cts_nomes[index] and (ct.id not in ids_used):
                    id_ct=ct.id
                    break

            if id_ct!=-1:

                ids_used.append(id_ct)

                for index_semana in range(0,53):

                    n_turnos = turnos[index]

                    for index_excecao in range(len(ct_excecao)):
                        if ct_excecao[index_excecao] == cts_nomes[index] and cc_excecao[index_excecao] == areas[
                            index] and index_semana >= de_excecao[index_excecao] and index_semana <= ate_excecao[
                            index_excecao]:
                            n_turnos = n_turnos_excecao[index_excecao]
                            break

                    try:
                        capacidade = int(dias_semana[index_semana]) * n_turnos * oee[index] * 60 * horas_por_turno[
                            index]
                    except:
                        capacidade = 0

                    if id_ct!=-1:
                        cts[id_ct].capacidade.append(capacidade)
                        cts[id_ct].capacidade_virados.append(capacidade_vir)
                        cts[id_ct].capacidade_blocos.append(capacidade_bl)
                        cts[id_ct].capacidade_iniciais.append(capacidade)

                        try:

                            cts[id_ct].capacidade[-1]=cts[id_ct].capacidade[-1]-ocupacoes[index_semana][index]*capacidade

                        except:
                            cts[id_ct].capacidade[-1] = cts[id_ct].capacidade[-1]

                cts[id_ct].capacidade_clientes=[]
                cts[id_ct].capacidade_iniciais_clientes=[]
                for posicao in range(len(clientes)):
                    capacidade_cliente_atual = [0]*len(cts[id_ct].capacidade_iniciais)
                    cts[id_ct].capacidade_clientes.append(capacidade_cliente_atual)
                    cts[id_ct].capacidade_iniciais_clientes.append(capacidade_cliente_atual)

                if cts_nomes[index] in ct_clientes:

                    posicao_clientes=find(ct_clientes,cts_nomes[index])

                    for id_posicao in range(len(posicao_clientes)):

                        posicao=posicao_clientes[id_posicao]

                        grupo_tofind=grupo_cliente[posicao]
                        capacidade=capacidade_cliente[posicao]

                        for id_cliente in range(len(clientes)):

                            if clientes[id_cliente].nome==grupo_tofind:

                                capacidade_cliente_atual = [i * capacidade for i in cts[id_ct].capacidade_iniciais]

                                cts[id_ct].capacidade_clientes[id_cliente]=capacidade_cliente_atual
                                cts[id_ct].capacidade_iniciais_clientes[id_cliente] = capacidade_cliente_atual

                if capacidade>0:
                    counter+=1
                helper.append(str(cts_nomes[index]))

        pbar.update(1)

    pbar.close()

    return cts

cts=[]
global semana_inicio_plano
semana_inicio_plano = datetime.datetime.now().isocalendar()[1]
if method == 0:
    semana_inicio_plano = semana_inicio_plano + 1
else:
    semana_inicio_plano = semana_inicio_plano + 4

df_blocos = pd.read_csv('data/04. restricoes blocos.csv', sep=",", encoding="iso-8859-1",error_bad_lines=False)
df_virados = pd.read_csv('data/05. restricoes virados.csv', sep=",", encoding="iso-8859-1",error_bad_lines=False)
df_clientes = pd.read_csv('data/06. restricoes clientes.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
cts=importar_cts('CNM_Planeamento_SAP_2022.xlsx',df_blocos,df_virados,df_clientes,cts,"CNM",method)
cts=importar_cts('CRM_Planeamento_SAP_2022.xlsx',df_blocos,df_virados,df_clientes,cts,"CRM",method)
cts=importar_cts_ccs('CCS_Planeamento_SAP_2022.xlsx',df_blocos,df_virados,df_clientes,cts,"CCS",method)

export_capacidades_in()

def print_capacidades():

    rows=[]
    global semana_inicio_plano

    for index in range(len(cts)):
        for semana in range(len(cts[index].capacidade_iniciais)):
            new_row={'Centro de Trabalho':cts[index].nome,'Acabamento':cts[index].acabamento, 'semana': semana + semana_inicio_plano, 'capacidade':cts[index].capacidade_iniciais[semana] }
            rows.append(new_row)

    df=pd.DataFrame(rows)
    df.to_csv('data/12. capacidade inicial.csv')

def read_arguments():
    global correr_cumprimento

    arguments=pd.read_csv('data/arguments.csv',error_bad_lines=False)

    cumprimento=arguments['correr_plano'].tolist()[0]

    return cumprimento

correr_cumprimento=read_arguments()

def import_ofs(method):

    lista_materials=ler_cts_embalagem()

    read_file= pd.read_excel("ACC_Ordens por planear - Kaizen.xlsx") #todo testar só com read_excel
    read_file.to_csv("data/08.  ofs.csv", index=None, header=True, date_format='%d/%m/%Y', encoding='iso-8859-1',errors='ignore')

    ofs=[]

    df_ofs = pd.read_csv('data/08.  ofs.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)

    df_tipo=pd.read_csv('data/15. marc.csv', sep=",",encoding='iso-8859-1',error_bad_lines=False)

    df_tipo=df_tipo[pd.to_numeric(df_tipo['Material'], errors='coerce').notnull()]
    material_ordem=df_tipo['Material'].tolist()
    tipo_de_ordem=df_tipo['Perfil de controle de produção'].tolist()

    df_ofs=df_ofs.drop_duplicates()
    df_ofs = df_ofs[df_ofs['Ordem de produção / planeada'].notnull()]
    df_ofs['Ordem de produção / planeada']=df_ofs['Ordem de produção / planeada'].astype(int)
    df_ofs['Qtd.necessária']=df_ofs['Qtd.necessária'].astype(str)
    df_ofs['Qtd.necessária'] = df_ofs['Qtd.necessária'].str.replace(' ', '')

    df_ofs['Qtd.necessária'] = df_ofs['Qtd.necessária'].astype(float)
    df_ofs['Duração da operação'] = df_ofs['Duração da operação'].astype(str)
    df_ofs['Duração da operação'] = df_ofs['Duração da operação'].str.replace(' ', '')
    df_ofs['Duração da operação'] = df_ofs['Duração da operação'].astype(float)

    df_ofs['Quantidade total da ordem'] = df_ofs['Quantidade total da ordem'].astype(str)
    df_ofs['Quantidade total da ordem']=df_ofs['Quantidade total da ordem'].str.replace(' ','')
    df_ofs['Quantidade total da ordem']=df_ofs['Quantidade total da ordem'].astype(float)

    if method == 0:  # por planear

        df_ofs = df_ofs[df_ofs['Ordem Venda / Transferência'].notna()]

        ovs_com_planeadas=df_ofs[(df_ofs['Ordem de produção / planeada'] > 1600000000)]
        df_ofs = df_ofs[(df_ofs['Ordem de produção / planeada'] < 400000000) & (df_ofs['Ordem de produção / planeada'] > 300000000)]

        ovs_com_planeadas['data_substituir']=ovs_com_planeadas['Data-base do fim']
        ovs_com_planeadas=ovs_com_planeadas[['Ordem Venda / Transferência','data_substituir']]
        ovs_com_planeadas=ovs_com_planeadas.groupby(['Ordem Venda / Transferência']).max().reset_index()
        df_ofs=df_ofs.merge(ovs_com_planeadas,on='Ordem Venda / Transferência',how='left')
        df_ofs['data_substituir']=df_ofs['data_substituir'].fillna(df_ofs['Data-base do fim'])
        df_ofs['Data-base do fim']=df_ofs['data_substituir']
        ovs_com_planeadas['data_substituir'] = pd.to_datetime(ovs_com_planeadas['data_substituir'], format='%d/%m/%Y',
                                                              errors='coerce')
        ovs_com_planeadas['semana'] = ovs_com_planeadas['data_substituir'].apply(lambda x: x.isocalendar()[1])
        ovs_com_planeadas.to_csv('data/111. ovs com ofs planeadas.csv')

    elif method==1: # só planeadas
        df_ofs = df_ofs[(df_ofs['Ordem de produção / planeada'] > 1600000000)]

    df_ofs['planeador']=df_ofs.iloc[:,0].str.split(' ').str[0]
    df_ofs['acabamento']=df_ofs['Texto breve de material'].str.split('/').str[1]
    df_ofs['acabamento'] = df_ofs['acabamento'].str.split(' ').str[0]
    df_ofs['descricao_bloco']=df_ofs['Descritivo componente'].str.split(' ').str[0]

    # limpar os que não têm o ct para depois verificar o que é preciso encomendar

    # criar lista de nomes dos centros de trabalho

    df_ofs['encomendar'] = df_ofs['Descritivo componente'].str.split(' ').str[-1]

    df_ofs = df_ofs.sort_values(['Ordem Venda / Transferência', 'Item OV/Transferência', 'Sold to'], ascending=[False, False, False])

    df_ofs['Data-base do fim'] = pd.to_datetime(df_ofs['Data-base do fim'], format='%d/%m/%Y')

    df_ofs['Data-base do fim']=df_ofs['Data-base do fim'].fillna(datetime.datetime.now())

    lista_ovs=df_ofs['Ordem Venda / Transferência'].tolist()
    cod_ofs = df_ofs['Ordem de produção / planeada'].tolist()
    lista_items = df_ofs['Item OV/Transferência'].tolist()

    cod_ov_in=1000

    for index in range(len(cod_ofs)):
        if math.isnan(lista_ovs[index]):
            lista_ovs[index]=cod_ov_in
            cod_ov_in+=1

        if math.isnan(lista_items[index]):
            lista_items[index]=99

    df_ofs['ov'] = lista_ovs
    df_ofs['item']=lista_items

    #remover todas as ovs que têm pelo menos um material NE

    if method==0:

        df_to_remove=df_ofs.copy()
        df_to_remove=df_to_remove[df_to_remove['encomendar']!='NE']
        df_to_remove=df_to_remove[['ov','item']]
        df_to_remove=df_to_remove.drop_duplicates()

        ovs_to_remove=df_to_remove['ov'].tolist()
        items_to_remove=df_to_remove['item'].tolist()

        df_to_remove=df_to_remove[df_to_remove['ov']>3000000000]
        df_to_remove=df_to_remove.merge(df_ofs,on=['ov','item'])

        df_to_remove['planeador']=np.where(df_to_remove['Descritivo de Centro de Trabalho'].str.contains('EMBALAGEM'),df_to_remove['Nome planejador'],"")
        df_to_remove=df_to_remove[df_to_remove['planeador']!=""]

        df_to_remove.to_csv('data/109. ovs removidas.csv')

        df_ofs = df_ofs[(~df_ofs['ov'].isin(ovs_to_remove))|(~df_ofs['item'].isin(items_to_remove))]

    df_ofs=df_ofs.sort_values(['ov', 'item','Material de produção'], ascending=[False, False,False])

    df_ofs['key'] = df_ofs['ov'].astype(str) + ' ' + df_ofs['item'].astype(str)

    df_cliente_final=df_ofs.copy()
    df_cliente_final=df_cliente_final[['ov', 'item','Sold to','planeador','Data-base do fim','Data de criação da ov']]
    df_cliente_final['key']=df_cliente_final['ov'].astype(str) + ' ' + df_cliente_final['item'].astype(str)
    df_cliente_final = df_cliente_final.drop_duplicates(subset=['key'], keep='first')
    df_cliente_final=df_cliente_final[['key','Sold to', 'planeador','Data-base do fim','Data de criação da ov']]
    df_cliente_final=df_cliente_final.sort_values(['key','planeador'],ascending=True)
    df_cliente_final=df_cliente_final.drop_duplicates(['key'])

    df_cliente_final['Data-base do fim'] = df_cliente_final[
        'Data-base do fim'].fillna(datetime.datetime.now())

    df_ofs=df_ofs.merge(df_cliente_final,how='left',on='key')
    df_ofs['semana'] = df_ofs.apply(lambda row: row['Data-base do fim_y'].isocalendar()[1], axis=1)
    df_ofs['semana']=df_ofs['semana']-semana_inicio_plano-1

    df_ofs = df_ofs.rename(columns={'Data-base do fim_y': 'Data-base do fim'})
    df_ofs = df_ofs.rename(columns={'Data de criação da ov_x': 'Data de criação da ov'})

    df_semana_criacao=df_ofs.copy()
    df_semana_criacao=df_semana_criacao[['ov','Data de criação da ov']]
    df_semana_criacao = df_semana_criacao[df_semana_criacao['Data de criação da ov'].notna()]
    df_semana_criacao=df_semana_criacao.drop_duplicates(subset=['ov'], keep='last')

    df_semana_criacao['Data de criação da ov'] = pd.to_datetime(df_semana_criacao['Data de criação da ov'], format='%d/%m/%Y')

    df_semana_criacao['semana_criacao'] = df_semana_criacao.apply(lambda row: row['Data de criação da ov'].isocalendar()[1], axis=1)

    df_semana_criacao['monday1'] = df_semana_criacao.apply(lambda row: row['Data de criação da ov']- datetime.timedelta(days=row['Data de criação da ov'].weekday()),axis=1)
    df_semana_criacao['monday2'] = df_semana_criacao.apply(
        lambda row: datetime.datetime.today() - datetime.timedelta(days=datetime.date.today().weekday()),axis=1)
    df_semana_criacao['numero_semanas'] = ((pd.to_datetime(df_semana_criacao['monday2']) - pd.to_datetime(df_semana_criacao['monday1'])).dt.days)/7

    if method==0:
        df_output_por_planear=df_semana_criacao[df_semana_criacao['numero_semanas']>2]
        df_output_por_planear=df_output_por_planear.merge(df_ofs,on='ov',how='left')
        df_output_por_planear['planeador']=np.where((df_output_por_planear['Descritivo de Centro de Trabalho'].str.contains('EMBALAGEM'))|(df_output_por_planear['Material de produção'].isin(lista_materials)),df_output_por_planear['Nome planejador'],"")
        df_output_por_planear.to_csv('data/110. ovs criacao.csv')
        df_semana_criacao=df_semana_criacao[df_semana_criacao['numero_semanas']<=2]

    if method==1:
        df_ofs=df_ofs[df_ofs['semana']>=0]

    df_semana_criacao=df_semana_criacao[['ov','semana_criacao']]

    df_ofs=df_ofs.merge(df_semana_criacao,how='inner',on='ov')

    nomes_cts = []
    for ct in cts:
        nomes_cts.append(ct.nome)

    df_ofs = df_ofs[df_ofs['Centro de trabalho'].isin(nomes_cts)]

    #todo adicionar restrição para restantes metodos

    df_ovs=df_ofs.copy()
    df_ovs=df_ovs.sort_values(by=['ov','item','Sold to_x'])
    df_ovs['Sold to_x']=df_ovs['Sold to_x'].ffill()

    df_ovs = df_ovs.drop_duplicates(['ov','semana'])

    ovs_ov=df_ovs['ov'].tolist()
    id_cliente_interno=df_ovs['planeador_y'].tolist()
    id_cliente = df_ovs['Sold to_y'].tolist()
    data_desejada = df_ovs['semana'].tolist()
    semanas_criacao=df_ovs['semana_criacao'].tolist()
    sold_to=df_ovs['Sold to_x'].tolist()

    ovs = []

    print('importar ordens de venda')

    with tqdm(total=len(ovs_ov)) as pbar:

        for index in range(len(ovs_ov)):

            semana=data_desejada[index]
            cod_ov=ovs_ov[index]

            cliente=id_cliente[index]
            cliente_interno=id_cliente_interno[index]
            semana_criacao=semanas_criacao[index]

            id_cliente_ov=-1
            id_interno_ov=-1

            for id_lista in range(len(clientes)):
                lista_clientes=clientes[id_lista].descricao_clientes
                if cliente in lista_clientes:
                    id_cliente_ov=id_lista
                    break

            if cliente_interno=='CCS':
                for id_lista in range(len(clientes)):
                    if clientes[id_lista].nome=='cliente interno':
                        id_interno_ov = id_lista
                        break

            new_ov=ov(index,cod_ov,id_cliente_ov,id_interno_ov,semana,semana_criacao,sold_to[index])
            ovs.append(new_ov)
            pbar.update(1)

    pbar.close()

    df_items = df_ofs.copy()
    df_items = df_items.drop_duplicates(['ov','item','semana'])
    items=[]

    ovs_item = df_items['ov'].tolist()
    items_item = df_items['item'].tolist()
    semana_item=df_items['semana'].tolist()
    planeadores=df_items['Nome planejador'].tolist()

    for index in range(len(ovs_item)):

        cod_item=items_item[index]
        planeador=planeadores[index]
        for pos_ov in range(len(ovs)):
            if ovs_item[index]==ovs[pos_ov].cod_ov and semana_item[index]==ovs[pos_ov].data_desejada:
                id_ov=pos_ov
                ovs[pos_ov].id_items.append(index)

        new_item=item(index,cod_item,id_ov,planeador)
        items.append(new_item)

    df_ofs_unicas = df_ofs.drop_duplicates(['Ordem de produção / planeada'])
    df_cs=pd.read_csv('data/21. total ref c.csv',sep=",",encoding='ISO-8859-1',error_bad_lines=False)
    df_cs['codigo material']=df_cs['codigo material'].astype(str)
    materiais_c=df_cs['codigo material'].tolist()

    ops=df_ofs['Ordem de produção / planeada'].tolist()
    duracoes = df_ofs['Duração da operação'].tolist() * 60
    quantidades = df_ofs['Quantidade total da ordem'].tolist()
    codigos_material = df_ofs['Material de produção'].tolist()
    descricoes_material = df_ofs['Texto breve de material'].tolist()
    codigos_precedencia = df_ofs['Componente'].tolist()
    blocos = df_ofs['Qtd.necessária'].tolist()
    descricao_bloco = df_ofs['descricao_bloco'].tolist()
    data_alocada = df_ofs['Data-base do fim'].tolist()
    data_desejada = df_ofs['semana'].tolist()
    lista_items=df_ofs['item'].tolist()
    lista_ovs=df_ofs['ov'].tolist()
    df_acabamentos = pd.read_csv('data/07.  acabamentos.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
    lista_descricao_acabamento=df_acabamentos['acabamento'].tolist()
    lista_tipo_acabamento=df_acabamentos['tipo'].tolist()
    lista_cts=df_ofs['Centro de trabalho'].tolist()
    lista_acabamentos=df_ofs['acabamento'].tolist()
    lista_descricao_precedencia=df_ofs['Descritivo componente'].tolist()
    lista_quantidade_precedencia=df_ofs['Qtd.necessária'].tolist()
    planeadores=df_ofs['Nome planejador'].tolist()
    descritivos_ct=df_ofs['Descritivo de Centro de Trabalho'].tolist()

    df_estufa = pd.read_csv('data/02. estufa.csv', skiprows=2, encoding='iso-8859-1',error_bad_lines=False)
    df_estufa['material'] = df_estufa['PRODUTO']

    df_estufa['TEMPO MINIMO\n(horas)'] = df_estufa['TEMPO MINIMO\n(horas)'].astype(float).fillna(0.0)
    df_estufa['duracao_mins'] = df_estufa['TEMPO MINIMO\n(horas)'] * 60

    estufa_material = df_estufa['material'].tolist()
    estufa_tempo = df_estufa['duracao_mins'].tolist()

    df=pd.read_csv('data/24. tintas.csv',sep=",",encoding="ISO-8859-1",error_bad_lines=False)
    df_verde=df[df['Revestimento']=="Tinta verde"]
    df_azul = df[df['Revestimento'] == "Tinta azul"]

    acabamento_verde=df_verde['Código'].tolist()
    acabamento_azul = df_azul['Código'].tolist()

    #todo adicionar drop duplicates

    pbar.close()

    ofs=[]

    print('importar ordens de fabrico')

    with tqdm(total=len(ops)) as pbar:

        index=0

        while index<(len(ops)):

            if index>0 and ops[index]==ops[index-1]:
                ofs[len(ofs)-1].codigo_precedencia.append(codigos_precedencia[index])

            else:

                cod_of=ops[index]
                duracao=duracoes[index]
                quantidade=quantidades[index]
                codigo_material=codigos_material[index]
                descricao_material=descricoes_material[index]
                viradas=0 #TODO adicionar viradas
                codigo_precedencia=codigos_precedencia[index]
                descricao_precedencia=lista_descricao_precedencia[index]
                quantidade_precedencia=lista_quantidade_precedencia[index]
                id_alocada=data_desejada[index]
                planeador=planeadores[index]
                descritivo_ct=descritivos_ct[index]

                tipo = ""
                for pos_material in range(len(material_ordem)):
                    if material_ordem[pos_material] == codigo_material:
                        tipo = tipo_de_ordem[pos_material]

                if descricao_bloco[index]=='BL':
                    n_blocos=float(blocos[index])
                else:
                    n_blocos=0

                if lista_acabamentos[index] in lista_descricao_acabamento and lista_cts[index][-2:]=='PL' and 'EMB' not in lista_cts[index]:
                    posicao=lista_descricao_acabamento.index(lista_acabamentos[index])
                    acabamento=lista_tipo_acabamento[posicao]
                elif descricao_material[:2] in lista_descricao_acabamento and lista_cts[index][:3]=='CCS':
                    posicao = lista_descricao_acabamento.index(lista_acabamentos[index])
                    if lista_tipo_acabamento[posicao]==lista_cts[index]:
                        acabamento = lista_tipo_acabamento[posicao]
                    else:
                        acabamento = 'Total'
                else:
                    acabamento='Total'

                if acabamento!='E01' and lista_cts[index][:3]=='CNM':
                    acabamento='Total'

                id_ct = -1

                for pos_ct in range(len(cts)):

                    if cts[pos_ct].nome == lista_cts[index] and cts[pos_ct].acabamento == acabamento:
                        id_ct = pos_ct
                        break
                    elif cts[pos_ct].nome==lista_cts[index] and lista_cts[index]=="CCSEMBME":
                        if descricoes_material[index].split(" ")[0]=="PL" and cts[pos_ct].acabamento=="Embalagem Placas":
                            id_ct = pos_ct
                            break
                        elif descricoes_material[index].split(" ")[0]=="MB" and cts[pos_ct].acabamento=="Embalagem Memos":
                            id_ct = pos_ct
                            break
                        elif (descricoes_material[index].split(" ")[0]!="MB" and descricoes_material[index].split(" ")[0]!="PL" ) and cts[pos_ct].acabamento=="Memoboards Diversos  (H691)":
                            id_ct = pos_ct
                            break
                    elif cts[pos_ct].nome==lista_cts[index] and lista_cts[index]=="CCSBALJT":
                        if descricoes_material[index].split(" ")[0]=="JT" and cts[pos_ct].acabamento=="Juntas":
                            id_ct = pos_ct
                            break
                        elif descricoes_material[index].split(" ")[0]!="JT" and cts[pos_ct].acabamento=="Outros":
                            id_ct = pos_ct
                            break

                if id_ct!=-1:

                    id_items = []

                    for pos_item in range(len(items)):

                        id_ov = items[pos_item].id_ov

                        if lista_items[index] == items[pos_item].cod_item and lista_ovs[index]==ovs[id_ov].cod_ov and data_desejada[index]==ovs[id_ov].data_desejada:
                            id_items.append(pos_item)
                            items[pos_item].id_ofs.append(len(ofs))

                    new_of = of(len(ofs), cod_of, duracao * 60, quantidade, id_ct, codigo_material, descricao_material,
                                codigo_precedencia,
                                id_items, n_blocos, viradas,descricao_precedencia,quantidade_precedencia,tipo,planeador,descritivo_ct)
                    if new_of.codigo_material in materiais_c:
                        new_of.is_c=1

                    if cts[new_of.id_ct].nome=="CRMLAMRL":
                        if new_of.acabamento in acabamento_verde:
                            new_of.tinta=1
                        elif new_of.acabamento in acabamento_azul:
                            new_of.tinta=0

                    ofs.append(new_of)

                    if method==1:
                        new_of.id_alocada.append(id_alocada)
                        new_of.alocada_duracao.append(duracao * 60)

                    if new_of.cod_material in estufa_material and cts[id_ct].nome == 'CRMLAMRL':
                        for posicao_codigo in range(len(estufa_material)):
                            if estufa_material[posicao_codigo] == new_of.cod_material:
                                tempo_estufa = estufa_tempo[posicao_codigo]
                                break

                        for posicao_ct in range(len(cts)):
                            if cts[posicao_ct].nome == 'CRMESTUFA':
                                id_ct_estufa = cts[posicao_ct].id
                                descritivo_ct_estufa = "CRMESTUFA"
                                break

                        # of_precedente = of(count, 40000000.0 + count,quantidade_precedencia ,quantidade_precedencia, id_ct_estufa,
                        #                    800000 + count, str(descricao_material + " ESTUFA"), codigo_precedencia, id_items, n_blocos, viradas,descricao_precedencia,quantidade_precedencia,tipo,planeador,
                        #                    descritivo_ct_estufa)
                        # ofs.append(of_precedente)
                        # new_of.codigo_precedencia.append(800000 + count)
                        # new_of.descricao_precedencia = str(descricao_material + " ESTUFA")
                        for id_item in id_items:
                            items[id_item].id_ofs.append(len(ofs))


            index+=1

            pbar.update(1)

    pbar.close()

    return ovs,items,ofs

ovs,items,ofs=import_ofs(method)

def import_bom():

    global lista_calandrados
    global lista_plyups

    df_bom=pd.read_csv('data/19. BOM.csv',encoding='iso-8859-1',error_bad_lines=False)

    df_calandrados=df_bom[(df_bom['Descrição Operação'].str.contains("CALANDRADOS")) & (df_bom['Centro trabalho'].str[:3]=='CRM')]
    df_plyups = df_bom[(df_bom['Descrição Operação'].str.contains("PLY-UP")) & (df_bom['Centro trabalho'].str[:3] == 'CRM')]

    df_calandrados['Material'].astype(int)
    df_plyups['Material'].astype(int)

    lista_calandrados=list(map(float,df_calandrados['Material'].tolist()))
    lista_plyups = list(map(float,df_plyups['Material'].tolist()))

    for pos_of in range(len(ofs)):

        id_of=ofs[pos_of].id

        try:
            precedencias=list(map(float,ofs[id_of].codigo_precedencia))
            if any(x in precedencias for x in lista_calandrados):
                ofs[id_of].calandrado = 1
            elif any(x in precedencias for x in lista_plyups):
                ofs[id_of].plyup = 1

        except:
            precedencias=""



    return lista_calandrados,lista_plyups

def atualizar_capacidades(method):
    df_estufa = pd.read_csv('data/02. estufa.csv', skiprows=2, encoding='iso-8859-1',error_bad_lines=False)
    df_estufa['material'] = df_estufa['PRODUTO']

    df_estufa['TEMPO MINIMO\n(horas)'] = df_estufa['TEMPO MINIMO\n(horas)'].astype(float).fillna(0.0)
    df_estufa['duracao_mins'] = df_estufa['TEMPO MINIMO\n(horas)'] * 60

    estufa_material = df_estufa['material'].tolist()
    estufa_tempo = df_estufa['duracao_mins'].tolist()

    if method==0 or method==1:

        df_ofs = pd.read_csv('data/08.  ofs.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
        df_acabamentos = pd.read_csv('data/07.  acabamentos.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
        lista_descricao_acabamento = df_acabamentos['acabamento'].tolist()
        lista_tipo_acabamento = df_acabamentos['tipo'].tolist()

        df_ofs.drop_duplicates(['Ordem de produção / planeada'])

        df_ofs = df_ofs[df_ofs['Ordem de produção / planeada'].notnull()]

        df_ofs['Ordem de produção / planeada'] = df_ofs['Ordem de produção / planeada'].astype(int)
        # df_ofs=df_ofs[df_ofs['Tipo de Ordem de produção/planeada']!='KD']
        df_ofs = df_ofs[(df_ofs['Ordem de produção / planeada'] > 1600000000)]

        df_ofs['Data-base do fim'] = df_ofs['Data-base do fim'].dropna()

        df_ofs['acabamento'] = df_ofs['Texto breve de material'].str.split('/').str[1]
        df_ofs['acabamento'] = df_ofs['acabamento'].str.split(' ').str[0]

        df_ofs['Data-base do fim'] = pd.to_datetime(df_ofs['Data-base do fim'], format='%d/%m/%Y')

        df_ofs=df_ofs[pd.notnull(df_ofs['Data-base do fim'])]

        df_ofs['semana'] = df_ofs.apply(lambda row: row['Data-base do fim'].isocalendar()[1],axis=1)

        df_ofs['Duração da operação']=df_ofs['Duração da operação'].astype(str)
        df_ofs['Duração da operação'] = df_ofs['Duração da operação'].str.replace(' ', '')
        df_ofs['Duração da operação'] = df_ofs['Duração da operação'].astype(float)

        #TODO ADICIONAR RESTRIÇAO CAPACIDADE CLIENTES,BLOCOS E VIRADAS

        df_ofs=df_ofs[df_ofs['semana']>=semana_inicio_plano]

        #SEMANA EM QUE SE RETIRA CAPACIDADE

        df_ofs['semana']=df_ofs['semana']-semana_inicio_plano

        df_ofs=df_ofs[['Centro de trabalho','semana','Duração da operação','acabamento','Ordem de produção / planeada','Ordem Venda / Transferência','Item OV/Transferência','Material de produção','Texto breve de material','Quantidade total da ordem','Quantidade fornecida ordem de produção']]

        df_ofs=df_ofs.reset_index()

        lista_cts=df_ofs['Centro de trabalho'].tolist()
        lista_semanas=df_ofs['semana'].tolist()
        lista_duracao=df_ofs['Duração da operação'].tolist()
        lista_acabamentos=df_ofs['acabamento'].tolist()
        ordem=df_ofs['Ordem de produção / planeada'].tolist()
        ov=df_ofs['Ordem Venda / Transferência'].tolist()
        item=df_ofs['Item OV/Transferência'].tolist()
        material_producao=df_ofs['Material de produção'].tolist()
        descricao_material=df_ofs['Texto breve de material'].tolist()
        quantidade_total=df_ofs['Quantidade total da ordem'].tolist()
        quantidade_fornecida=df_ofs['Quantidade fornecida ordem de produção'].tolist()
        codigos_material = df_ofs['Material de produção'].tolist()

        total=0
        rows=[]

        with tqdm(total=len(lista_cts)) as pbar:

            for index in range(len(lista_duracao)):

                if lista_acabamentos[index] in lista_descricao_acabamento and lista_cts[index][-2:]=='PL':
                    posicao=lista_descricao_acabamento.index(lista_acabamentos[index])
                    acabamento=lista_tipo_acabamento[posicao]
                elif descricao_material[:2] in lista_descricao_acabamento and lista_cts[index][:3]=='CCS':
                    posicao = lista_descricao_acabamento.index(lista_acabamentos[index])
                    if lista_tipo_acabamento[posicao]==lista_cts[index]:
                        acabamento = lista_tipo_acabamento[posicao]
                    else:
                        acabamento = 'Total'
                else:
                    acabamento='Total'

                # if material_producao[index] in material_c:
                #     aglomerado=aglomerado_c[material_c.index(material_producao[index])]
                # else:
                #     aglomerado=-1

                aglomerado = -1

                if acabamento!='Total':
                    new_row = {'Semana': lista_semanas[index] + semana_inicio_plano,
                               'Centro de Trabalho': lista_cts[index], 'Acabamento': 'Total', 'OV': ov[index],
                               'Item': item[index], 'OF': ordem[index], 'Material': material_producao[index],
                               'Descrição': descricao_material[index], 'Quantidade': quantidade_total[index],
                               'Quantidade Fornecida': quantidade_fornecida[index], 'aglomerado': aglomerado}
                    rows.append(new_row)
                new_row={'Semana':lista_semanas[index]+semana_inicio_plano,'Centro de Trabalho':lista_cts[index],'Acabamento':acabamento,'OV':ov[index],'Item':item[index],'OF':ordem[index],'Material':material_producao[index],'Descrição':descricao_material[index],'Quantidade':quantidade_total[index],'Quantidade Fornecida':quantidade_fornecida[index],'aglomerado':aglomerado}
                rows.append(new_row)
                if codigos_material[index] in estufa_material and lista_cts[index] == 'CRMLAMRL':
                    new_row = {'Semana': lista_semanas[index] + semana_inicio_plano,
                               'Centro de Trabalho': 'CRMESTUFA', 'Acabamento': 'Total', 'OV': ov[index],
                               'Item': item[index], 'OF': ordem}
                    rows.append(new_row)


                pbar.update(1)
        pbar.close()

        df=pd.DataFrame(rows)
        df.to_csv('data/ofs_planeadas.csv')

atualizar_capacidades(method)

end=time.time()

def print_capacidade_reservada(method):

    rows=[]

    for id_ct in range(len(cts)):

        for semana in range(len(cts[id_ct].capacidade)):

            if method==0 or method==1:

                new_row = {'centro de trabalho': cts[id_ct].nome, 'acabamento': cts[id_ct].acabamento,
                           'semana': semana + semana_inicio_plano,
                           'carga ocupada': cts[id_ct].capacidade_iniciais[semana] - cts[id_ct].capacidade[semana]}

                rows.append(new_row)

            else:

                new_row = {'centro de trabalho': cts[id_ct].nome, 'acabamento': cts[id_ct].acabamento,
                           'semana': semana + semana_inicio_plano,
                           'carga ocupada': 0}

                rows.append(new_row)

    df = pd.DataFrame(rows)
    df.to_csv('data/11. ocupacao ofs planeadas.csv')

def sort_by_leadtime():
    global ovs
    global clientes

    if method==0:

        sorted_index=[]
        sorted_leadtimes=[]

        for index in range(len(ovs)):

            id_cliente=int(ovs[index].id_cliente)

            leadtime_cliente=clientes[id_cliente].lead_time

            semana_desejada=ovs[index].data_desejada

            semana_criacao=ovs[index].data_criacao

            if leadtime_cliente!=0 and id_cliente!=-1:

                data_com_leadtime = semana_criacao + leadtime_cliente

                data_min=min(semana_desejada,data_com_leadtime)

                ovs[index].data_min=data_min

            else:

                data_min=semana_desejada

            posicao_alocar = bisect.bisect_left(sorted_leadtimes, data_min)
            bisect.insort(sorted_leadtimes, data_min)
            sorted_index.insert(posicao_alocar,index)

    else:

        sorted_index = []
        sorted_leadtimes = []

        for index in range(len(ovs)):

            semana_max=0

            id_cliente = int(ovs[index].id_cliente)

            leadtime_cliente = clientes[id_cliente].lead_time

            semana_desejada = ovs[index].data_desejada

            semana_criacao = ovs[index].data_criacao

            for pos_item in range(len(ovs[index].id_items)):

                id_item=ovs[index].id_items[pos_item]

                for pos_of in range(len(items[id_item].id_ofs)):

                    id_of=items[id_item].id_ofs[pos_of]

                    if len(ofs[id_of].id_alocada)>0:

                        if max(ofs[id_of].id_alocada)>semana_max:
                            semana_max=max(ofs[id_of].id_alocada)

            if semana_max>=0:

                if leadtime_cliente != 0 and id_cliente != -1:

                    data_com_leadtime = semana_criacao + leadtime_cliente

                    data_min = min(semana_desejada, data_com_leadtime)

                    ovs[index].data_min = data_min

                else:

                    data_min = semana_desejada

                posicao_alocar = bisect.bisect_left(sorted_leadtimes, data_min)
                bisect.insort(sorted_leadtimes, data_min)
                sorted_index.insert(posicao_alocar, index)


    return sorted_index

def verificar_capacidades(id_of,int_semana,id_ov,capacidade_max,duracao,lista_verdes,lista_azuis):

    global ofs
    global cts
    global ovs

    id_ct = ofs[id_of].id_ct

    resultado=True

    #verificar blocos
    blocos=ofs[id_of].blocos
    viradas = ofs[id_of].viradas
    cliente_interno = ovs[id_ov].id_interno
    cliente_final = ovs[id_ov].id_cliente

    if ofs[id_of].calandrado == 1:
        max_capacidade = max(capacidade_calandrados)
        count_refs = 0
        pos_material = -1
        material_atual = -1

        for index in range(len(pos_calandrados)):
            if pos_calandrados[index]==int(ofs[id_of].codigo_material):
                material_atual=index
                break

        # for sublist in capacidade_calandrados:
        #     pos_material += 1
        #     if sublist[int_semana] != max_capacidade and pos_material!=material_atual:
        #         count_refs += 1

        if count_refs>=2:
            return False

        if capacidade_calandrados[material_atual][int_semana]+ofs[id_of].quantidade_precedencia<0:
            return False

    if ofs[id_of].plyup == 1:
        max_capacidade = max(capacidade_plyups)
        max_capacidade=max(max_capacidade)
        count_refs = 0
        pos_material = -1
        material_atual=-1

        for index in range(len(pos_plyups)):
            if pos_plyups[index]==int(ofs[id_of].codigo_material):
                material_atual=index
                break

        for sublist in capacidade_plyups:
            pos_material += 1
            if sublist[int_semana] != max_capacidade and pos_material!=material_atual:
                count_refs += 1

        if count_refs>=2:
            return False

        if material_atual!=-1:
            if capacidade_plyups[material_atual][int_semana]+ofs[id_of].quantidade_precedencia<0:
                return False

    if cts[id_ct].capacidade_blocos[int_semana]<blocos:
        resultado=False

    #verificar viradas

    if cts[id_ct].capacidade_virados[int_semana]<viradas:
        resultado=False

    #verificar cliente interno

    if cliente_interno!=-1:
        print('ct: ' + str(id_ct) + ' semana: ' + str(int_semana) + ' cliente ' + str(cliente_interno))
        if cts[id_ct].capacidade_clientes[cliente_interno][int_semana]<duracao and cts[id_ct].capacidade_iniciais_clientes[cliente_interno][int_semana]>0:
            resultado=False
        elif (int_semana<2 and cts[id_ct].capacidade[int_semana] - duracao < (1-capacidade_max)*cts[id_ct].capacidade_iniciais[int_semana]) or cts[id_ct].capacidade[int_semana]<duracao:
            resultado=False
        elif int_semana>2 and cts[id_ct].capacidade[int_semana]-cts[id_ct].capacidade_clientes[cliente_interno][int_semana]-duracao<(1-capacidade_max)*cts[id_ct].capacidade_iniciais[int_semana] and cts[id_ct].capacidade[int_semana]<duracao:
            resultado=False

    # verificar cliente final

    if cliente_final!=-1:
        if cts[id_ct].capacidade_clientes[cliente_final][int_semana]<duracao and cts[id_ct].capacidade_iniciais_clientes[cliente_final][int_semana]>0:
            resultado=False

    # verificar capacidade

    if (int_semana< 2 and cts[id_ct].capacidade[int_semana] - duracao < (1-capacidade_max)*cts[id_ct].capacidade_iniciais[int_semana]):
        resultado=False

    if (int_semana>=2 and cts[id_ct].capacidade[int_semana] - duracao < (1-capacidade_max)*cts[id_ct].capacidade_iniciais[int_semana]):
        resultado=False

    if cts[id_ct].capacidade[int_semana]<duracao:
        resultado=False

    #setups tinta

    possivel_esta_semana=0
    possivel_proxima_semana=1

    for of in ofs:
        if int_semana in of.id_alocada and of.tinta==ofs[id_of].tinta and of.tinta!=-1:
            possivel_esta_semana=1
        elif int_semana+1 in of.id_alocada and of.tinta==ofs[id_of].tinta and of.tinta!=-1:
            possivel_proxima_semana=1
            resultado=False
        elif ofs[id_of].tinta==1 and lista_verdes[int_semana]==1 and of.tinta!=-1:
            possivel_esta_semana = 1
        elif ofs[id_of].tinta==1 and lista_verdes[int_semana+1]==1 and of.tinta!=-1:
            possivel_proxima_semana = 1
            resultado = False
        elif ofs[id_of].tinta==0 and lista_azuis[int_semana]==1 and of.tinta!=-1:
            possivel_esta_semana = 1
        elif ofs[id_of].tinta==0 and lista_azuis[int_semana+1]==1 and of.tinta!=-1:
            possivel_proxima_semana = 1
            resultado = False

    return resultado

def ler_tintas_alocadas():

    df=pd.read_csv('data/08.  ofs.csv',sep=",",encoding="ISO-8859-1",error_bad_lines=False)

    df=df[(df['Ordem de produção / planeada']>1600000000)&(df['Ordem de produção / planeada']<1700000000)]
    df['Data-base do fim'] = pd.to_datetime(df['Data-base do fim'], format='%d/%m/%Y',
                                                          errors='coerce')
    df=df[df['Centro de trabalho']=="CRMLAMRL"]
    last_week = date(datetime.datetime.today().year , 12, 28)
    n_weeks=last_week.isocalendar()[1]-semana_inicio_plano

    df['semana'] = df['Data-base do fim'].apply(lambda x: x.isocalendar()[1]-semana_inicio_plano if x.isocalendar()[0]==datetime.datetime.today().year else n_weeks+x.isocalendar()[1] )
    try:
        df=df[df['semana']>=0]
    except:
        x=0

    df_total = pd.read_csv('data/24. tintas.csv', sep=",", encoding="ISO-8859-1",error_bad_lines=False)
    df_verde = df_total[df_total['Revestimento'] == "Tinta verde"]
    df_azul = df_total[df_total['Revestimento'] == "Tinta azul"]

    acabamento_verde = df_verde['Código'].tolist()
    acabamento_azul = df_azul['Código'].tolist()

    descricao_material=df['Texto breve de material'].tolist()
    semana=df['semana'].tolist()

    lista_verdes=[-1]*(53*2)
    lista_azuis = [-1] * (53 * 2)

    for index in range(len(descricao_material)):

        semana_atual=semana[index]

        try:
            acabamento = descricao_material[index].split('/')[1].split(' ')[0]
        except:
            acabamento = ""

        if acabamento in acabamento_verde:
            lista_verdes[semana_atual]=1
        elif acabamento in acabamento_azul:
            lista_azuis[semana_atual]=1

    return lista_verdes,lista_azuis

def alocar(id_of,int_semana,id_ov,duracao):

    global ofs
    global cts
    global ovs

    id_ct = ofs[id_of].id_ct

    blocos = ofs[id_of].blocos
    viradas = ofs[id_of].viradas
    cliente_interno = ovs[id_ov].id_interno
    cliente_final = ovs[id_ov].id_cliente

    if ofs[id_of].duracao>0:
        fator_conversao=duracao/ofs[id_of].duracao
    else:
        fator_conversao=1

    material_atual=-1

    if ofs[id_of].calandrado == 1:

        pos_material = -1

        for index in range(len(pos_calandrados)):
            if pos_calandrados[index]==int(ofs[id_of].codigo_material):
                material_atual=index
                break

        capacidade_calandrados[material_atual][int_semana]=capacidade_calandrados[material_atual][int_semana]-fator_conversao*ofs[id_of].quantidade_precedencia

    if ofs[id_of].plyup == 1 and material_atual!=-1:

        pos_material = -1

        for index in range(len(pos_plyups)):
            if pos_plyups[index]==int(ofs[id_of].codigo_material):
                material_atual=index
                break

        capacidade_plyups[material_atual][int_semana]=capacidade_plyups[material_atual][int_semana]-fator_conversao*ofs[id_of].quantidade_precedencia

    cts[id_ct].capacidade_blocos[int_semana]=cts[id_ct].capacidade_blocos[int_semana]-blocos

    cts[id_ct].capacidade_virados[int_semana]=cts[id_ct].capacidade_virados[int_semana]-viradas

    if cliente_interno!=-1:
        cts[id_ct].capacidade_clientes[cliente_interno][int_semana] = cts[id_ct].capacidade_clientes[cliente_interno][int_semana]-duracao
    if cliente_final!=-1:
        cts[id_ct].capacidade_clientes[cliente_final][int_semana] = cts[id_ct].capacidade_clientes[cliente_final][int_semana] - duracao

    cts[id_ct].capacidade[int_semana]=cts[id_ct].capacidade[int_semana]-duracao

    if cts[id_ct].nome=='CNMLAMPL' and cts[id_ct].acabamento=='E01':

        for pos_ct in range(len(cts)):
            if cts[pos_ct].nome=='CNMLAMPL' and cts[pos_ct].acabamento=='Total':
                id_adicional=pos_ct
                cts[id_adicional].capacidade[int_semana]=cts[id_adicional].capacidade[int_semana]-duracao
                break

    ofs[id_of].id_alocada.append(int_semana)
    ofs[id_of].id_alocada_anterior.append(int_semana)
    ofs[id_of].alocada_duracao.append(duracao)
    ofs[id_of].alocada_duracao_anterior.append(duracao)

def desalocar(id_of,int_semana,id_ov,duracao):

    global ofs
    global cts
    global ovs

    id_ct = ofs[id_of].id_ct
    blocos = ofs[id_of].blocos
    viradas = ofs[id_of].viradas
    cliente_interno = ovs[id_ov].id_interno
    cliente_final = ovs[id_ov].id_cliente
    if ofs[id_of].duracao!=0:
        fator_conversao=duracao/ofs[id_of].duracao
    else:
        fator_conversao=0

    material_atual=-1

    cts[id_ct].capacidade_blocos[int_semana] = cts[id_ct].capacidade_blocos[int_semana] + blocos
    cts[id_ct].capacidade_virados[int_semana] = cts[id_ct].capacidade_virados[int_semana] + viradas

    if ofs[id_of].calandrado == 1 and material_atual!=-1:

        pos_material = -1

        for index in range(len(pos_calandrados)):
            if pos_calandrados[index] == int(ofs[id_of].codigo_material):
                material_atual = index
                break

        capacidade_calandrados[material_atual][int_semana] = capacidade_calandrados[material_atual][int_semana] + fator_conversao*ofs[
            id_of].quantidade_precedencia

    if ofs[id_of].plyup == 1 and material_atual!=-1:

        pos_material = -1

        for index in range(len(pos_plyups)):
            if pos_plyups[index] == int(ofs[id_of].codigo_material):
                material_atual = index
                break

        capacidade_plyups[material_atual][int_semana] = capacidade_plyups[material_atual][int_semana] + fator_conversao*ofs[
            id_of].quantidade_precedencia

    if cliente_interno != -1:
        cts[id_ct].capacidade_clientes[cliente_interno][int_semana] = cts[id_ct].capacidade_clientes[cliente_interno][int_semana] + duracao

    if cliente_final != -1:
        cts[id_ct].capacidade_clientes[cliente_final][int_semana] = cts[id_ct].capacidade_clientes[cliente_final][
                                                                        int_semana] + duracao

    cts[id_ct].capacidade[int_semana] = cts[id_ct].capacidade[int_semana] + duracao

    if cts[id_ct].nome=='CNMLAMPL' and cts[id_ct].acabamento=='E01':
        for pos_ct in range(len(cts)):
            if cts[pos_ct].nome=='CNMLAMPL' and cts[pos_ct].acabamento=='Total':
                id_adicional=pos_ct
                cts[id_adicional].capacidade[int_semana]=cts[id_adicional].capacidade[int_semana]+duracao
                break

def verificar_precedencias():

    global ofs
    global items

    for index in range(len(items)):

        id_item=items[index].id

        lista_ofs=items[id_item].id_ofs

        for pos_lista in range(len(lista_ofs)):

            id_of=lista_ofs[pos_lista]

            for pos_codigo in range(len(ofs[id_of].codigo_precedencia)):

                for pos_of_lista in range(len(lista_ofs)):

                    id_potencial=lista_ofs[pos_of_lista]
                    print(id_potencial)

                    if ofs[id_of].codigo_precedencia[pos_codigo]==ofs[id_potencial].codigo_material:

                        ofs[id_of].precedencias.append(id_potencial)
                        ofs[id_potencial].sequencias.append(id_of)

def verificar_possivel_atras(id_of,semana_pretendida,id_ov,semana_min_precedencia,lista_verdes,lista_azuis):

    global ofs
    global cts
    global ovs

    possivel=False
    count=semana_pretendida-1
    semana_possivel=-1

    while count>=semana_min_precedencia and possivel==False:
        possivel=verificar_capacidade_rolante(id_of,count,id_ov,0.5,0.95,lista_verdes,lista_azuis)
        if possivel == True:
            semana_possivel=count
        count-=1

    return semana_possivel

def minimizar_wip_item(id_item,max_semana,parametro_duracao):

    lista_ofs = items[id_item].id_ofs
    id_ov=items[id_item].id_ov

    if max_semana==-1: #aqui quero minimizar dentro do mesmo item, por isso tenho de calcular a max_semana real

        for index in range(len(lista_ofs)):

            id_of=lista_ofs[index]

            semana_alocada_max=max(ofs[id_of].id_alocada)

            if semana_alocada_max>max_semana:
                max_semana=semana_alocada_max

    for i in range(len(lista_ofs)-1,-1,-1):

        id_of=lista_ofs[i]

        of=ofs[id_of]

        semana_min_sucedencia=max_semana

        semana_min_precedencia = max_semana
        semana_max_precedencia = max_semana
        n_semanas_prec=1

        for pos_sucedente in range(len(of.sequencias)):
            id_sucedencia=of.sequencias[pos_sucedente]
            if min(ofs[id_sucedencia].id_alocada)<semana_min_sucedencia:
                semana_min_sucedencia=min(ofs[id_sucedencia].id_alocada)

        for pos_precedencia in range(len(of.precedencias)):
            id_precedencia=ofs[of.precedencias[pos_precedencia]].id
            if min(ofs[id_precedencia].id_alocada)<semana_min_precedencia:
                semana_min_precedencia=min(ofs[id_precedencia].id_alocada)
                n_semanas_prec>=ofs[id_precedencia].n_semanas

            if max(ofs[id_precedencia].id_alocada) < semana_max_precedencia:
                semana_max_precedencia = max(ofs[id_precedencia].id_alocada)

        #não queremos minimizar wip de tintas
        if max(ofs[id_of].id_alocada)!=max_semana and ofs[id_of].tinta==-1:
            #count = semana_max_precedencia-ofs[id_of].n_semanas+1
            #limite=semana_max_precedencia-ofs[id_of].n_semanas+1
            count = max_semana-ofs[id_of].n_semanas+1
            limite = semana_min_precedencia+n_semanas_prec-ofs[id_of].n_semanas
            duracao=ofs[id_of].duracao

            # if duracao==0:
            #     count=max_semana

            if count>semana_min_sucedencia and duracao!=0:
                count=semana_min_sucedencia-ofs[id_of].n_semanas+1

            alterada = False

            # if duracao==0:
            #     for semana_alocada in ofs[id_of].id_alocada:
            #         desalocar(id_of,semana_alocada,id_ov,duracao)
            #     ofs[id_of].id_alocada=[]
            #     alocar(id_of,max_semana,id_ov,duracao)
            #     alterada=True
            #
            # else:

            while alterada==False and count>=limite:

                id_ct=ofs[id_of].id_ct

                if (verificar_capacidade_rolante(id_of,count,id_ov,0.5,0.95,[],[])==True) and (verificar_acabamento(id_of,count,parametro_duracao)==True):

                    for index in range(len(ofs[id_of].id_alocada)):
                        semana_anterior=ofs[id_of].id_alocada[index]
                        duracao_anterior=ofs[id_of].alocada_duracao[index]
                        if id_ct!=-1:
                            desalocar(id_of,semana_anterior,id_ov,duracao_anterior)

                    ofs[id_of].id_alocada=[]
                    ofs[id_of].alocada_duracao=[]
                    alocar_capacidade_rolante(id_of, count, id_ov, 0.5, 0.95,[],[])

                    alterada=True

                count-=1

            if alterada==False:
                return limite


    return -1

def minimizar_wip(id_ov,parametro_duracao):

    max_semana=-1

    for index in range(len(ovs[id_ov].id_items)):
        n_ofs=len(items[ovs[id_ov].id_items[index]].id_ofs)
        id_ofs=items[ovs[id_ov].id_items[index]].id_ofs
        for i in range(len(id_ofs)):
            id_of=id_ofs[i]
            semana_alocada=max(ofs[id_of].id_alocada)
            if semana_alocada>max_semana:
                max_semana=semana_alocada

    id_items=ovs[id_ov].id_items

    for index in range(len(id_items)):

        id_item=id_items[index]
        minimizar_wip_item(id_item,max_semana,parametro_duracao)

    return -1

def print_output(method):

    global semana_inicio_plano

    rows=[]
    carga_planeada=[]

    for index in range(len(ofs)):

        id_of=ofs[index].id
        max_semana_precedencia=0

        for pos_precedencia in range(len(ofs[id_of].precedencias)):
            id_precedencia=ofs[id_of].precedencias[pos_precedencia]
            if len(ofs[id_precedencia].id_alocada)>0:
                if max(ofs[id_precedencia].id_alocada)>max_semana_precedencia:
                    max_semana_precedencia=max(ofs[id_precedencia].id_alocada)

        if ofs[index].id_ct!=-1 and len(ofs[index].id_alocada)!=0 and (-1 not in ofs[index].id_alocada):

            flag_setup=-1
            for pos_of in range(len(ofs)):
                if len(ofs[pos_of].id_alocada) > 0:
                    if ofs[pos_of].cod_material==ofs[index].cod_material and ofs[pos_of].dim1==ofs[index].dim1 and ofs[pos_of].dim2==ofs[index].dim2 and min(ofs[pos_of].id_alocada)<min(ofs[index].id_alocada) and ofs[index].duracao<=60:
                        if flag_setup==-1 or min(ofs[pos_of].id_alocada)<flag_setup:
                            flag_setup=min(ofs[pos_of].id_alocada)+semana_inicio_plano


            # if str(ofs[index].codigo_material) in material_c:
            #     aglomerado=aglomerado_c[material_c.index(str(ofs[index].codigo_material))]
            # else:
            aglomerado=-1

            semana_precedencia=max_semana_precedencia + semana_inicio_plano

            semana_output = max(ofs[index].id_alocada) + semana_inicio_plano

            new_row={'of':ofs[index].cod_of,
                     'item':items[ofs[index].id_items[0]].cod_item,
                     'id ov':items[ofs[index].id_items[0]].id_ov,
                     'ov':ovs[items[ofs[index].id_items[0]].id_ov].cod_ov,
                     'semana': max(ofs[index].id_alocada) + semana_inicio_plano,
                     'data desejada':ovs[items[ofs[index].id_items[0]].id_ov].data_desejada+semana_inicio_plano,
                     'centro de trabalho':cts[ofs[index].id_ct].nome,
                     'acabamento': cts[ofs[index].id_ct].acabamento,
                     'blocos':ofs[index].blocos,
                     'ccs':ovs[items[ofs[index].id_items[0]].id_ov].id_interno,
                     'virados':ofs[index].viradas,
                     'referência':ofs[index].descricao_material,
                     'semana minima':ofs[index].semana_min+semana_inicio_plano,
                     'duração':ofs[index].duracao,
                     'semana precedencia':semana_precedencia,
                     'número de semanas':ofs[index].n_semanas,
                     'Código Material':ofs[index].codigo_material,
                     'Quantidade Material':ofs[index].quantidade,
                     'Código Consumo':ofs[index].codigo_precedencia[0],
                     'Quantidade Consumo':ofs[index].quantidade_precedencia,
                     'Descrição Consumo':ofs[index].descricao_precedencia,
                     'Semana com mesma ref':flag_setup,
                     'aglomerado':aglomerado,
                    'c':ofs[index].is_c}

            for pos_alocada in range(len(ofs[index].id_alocada)):
                new_carga={'semana':ofs[index].id_alocada[pos_alocada]+semana_inicio_plano,
                                'duracao':ofs[index].alocada_duracao[pos_alocada],
                                'centro de trabalho':cts[ofs[index].id_ct].nome,
                                'acabamento':cts[ofs[index].id_ct].acabamento}
                carga_planeada.append(new_carga)

            rows.append(new_row)

    df = pd.DataFrame(rows)
    df_cts=pd.DataFrame(carga_planeada)

    try:
        df_cts=df_cts.groupby(by=['centro de trabalho','acabamento','semana']).sum()
        df_cts=df_cts.reset_index()
    except:
        print('Não foram planeadas OVs')

    if method==0:
        df.to_csv('data/09. output.csv')
        df_cts.to_csv('data/102. capacidade alocada.csv')
    elif method==1:
        df.to_csv('data/13. output planeado.csv')
    elif method==2:
        df.to_csv('data/14. output global.csv')

def print_mapa_precedencias():

    rows=[]

    for index in range(len(ofs)):

        lista_precedencias=ofs[index].precedencias

        lista_semanas=[]

        max_precedencia=-1

        for pos_precedencia in range(len(lista_precedencias)):

            id_precedencia=lista_precedencias[pos_precedencia]

            if len(ofs[id_precedencia].id_alocada)>0:

                semana_precedencia=min(ofs[id_precedencia].id_alocada)

                lista_semanas.append(semana_precedencia)

                max_precedencia=max(lista_semanas)

        if len(ofs[index].id_alocada)>0:
            semana_print=min(ofs[index].id_alocada)

        else:
            semana_print=-1

        new_row = {'of': ofs[index].cod_of, 'item': items[ofs[index].id_items[0]].cod_item,
                   'ov': ovs[items[ofs[index].id_items[0]].id_ov].cod_ov, 'semana': semana_print + semana_inicio_plano,
                   'semana precedencia': max_precedencia + semana_inicio_plano,
                   'numero de ofs':ofs[index].n_semanas}

        rows.append(new_row)

    df = pd.DataFrame(rows)
    df.to_csv('data/10. mapa precedencias.csv')

def partir_of(id_of):

    global ofs

    id_ct=ofs[id_of].id_ct

    count=0

    capacidades_iniciais=cts[id_ct].capacidade_iniciais

    capacidade_media=max(capacidades_iniciais)

    if ofs[id_of].duracao>0.5*capacidade_media:

        numero_ofs=ofs[id_of].duracao//(0.5*capacidade_media)

        if numero_ofs>1:

            count+=1

            nova_duracao=ofs[id_of].duracao/numero_ofs
            nova_quantidade=ofs[id_of].quantidade/numero_ofs
            novo_nblocos=ofs[id_of].blocos/numero_ofs
            novo_viradas=ofs[id_of].viradas/numero_ofs

            # alterar a duracao da primeira of

            ofs[id_of].duracao=nova_duracao
            ofs[id_of].nova_quantidade=nova_quantidade
            ofs[id_of].blocos=novo_nblocos
            ofs[id_of].viradas=novo_viradas


            # criar outras ofs com a nova duracao e com as precedencias da inicial

            id_nova=len(ofs)

            numero_ofs=int(numero_ofs)

            for pos in range(numero_ofs-1): #-1 porque já estamos a alterar a que existe

                id_precedencias=ofs[id_of].precedencias
                id_sequencias = ofs[id_of].sequencias

                new_of=of(len(ofs),ofs[id_of].cod_of,nova_duracao,nova_quantidade,id_ct,ofs[id_of].codigo_material,ofs[id_of].descricao_material,ofs[id_of].codigo_precedencia,ofs[id_of].id_items,novo_nblocos,novo_viradas)
                new_of.precedencias=id_precedencias
                new_of.sequencias=id_sequencias
                ofs.append(new_of)

                #adicionar ofs ao item:
                for pos_item in range(len(ofs[id_of].id_items)):
                    id_item=ofs[id_of].id_items[pos_item]

                    for pos_of_item in range(len(items[id_item].id_ofs)):

                        if items[id_item].id_ofs[pos_of_item]==id_of:

                            items[id_item].id_ofs.insert(pos_of_item,id_nova)
                            break

    return count

def verificar_acabamento(id_of,semana,duracao):

    if method==0:

        id_ct=ofs[id_of].id_ct

        if cts[id_ct].acabamento!='Total' and ofs[id_of].duracao<=duracao:

            if cts[id_ct].capacidade_iniciais[semana]!=cts[id_ct].capacidade[semana]:

                return True

            else:

                return False

        else:

            return True
    else:
        return True

def output_partidas():

    codigos=[]
    max=[]
    min=[]
    count=[]

    for index in range(len(ofs)):

        cod_of=ofs[index].cod_of
        semana_alocada=ofs[index].id_alocada

        if any(cod_of in sublist for sublist in codigos):

            for i, lst in enumerate(codigos):
                for j, color in enumerate(lst):
                    if color == cod_of:
                        posicao_codigo = i

            codigos[posicao_codigo].append(cod_of)

            count[posicao_codigo] = count[posicao_codigo] + 1

            if max[posicao_codigo]<semana_alocada:

                max[posicao_codigo]=semana_alocada

            if min[posicao_codigo]>semana_alocada:

                min[posicao_codigo]=semana_alocada

        else:

            codigos.insert(0,[cod_of])
            max.insert(0,semana_alocada)
            min.insert(0,semana_alocada)
            count.insert(0,1)

    rows=[]

    for index in range(len(count)):

        if count[index]>1:

            cod_of=codigos[index][0]
            max_semana=max[index]
            min_semana=min[index]

            new_row={'of':cod_of,'semana maxima':max_semana,'semana minima':min_semana,'numero ofs':count[index],'diferenca':max_semana-min_semana+1-count[index]}

            rows.append(new_row)

    df=pd.DataFrame(rows)

    df.to_csv('data/13. ofs partidas.csv')

def calcular_n_ofs(ocupacao_max_semana):

    for of in ofs:

        id_of=of.id

        duracao = of.duracao
        id_ct = of.id_ct
        capacidade_med = sum(cts[id_ct].capacidade_iniciais) / len(cts[id_ct].capacidade_iniciais)

        if of.duracao==0:
            n_semanas=1

        else:

            n_semanas = math.ceil(duracao / (ocupacao_max_semana * capacidade_med))

        of.n_semanas=n_semanas

def verificar_capacidade_rolante(id_of,int_semana,id_ov,ocupacao_max_semana,capacidade_max,lista_verdes,lista_azuis):

    duracao = ofs[id_of].duracao
    id_ct = ofs[id_of].id_ct

    n_semanas=ofs[id_of].n_semanas

    semana=int_semana

    capacidade_med = sum(cts[id_ct].capacidade_iniciais) / len(cts[id_ct].capacidade_iniciais)

    while duracao > 0 and semana <= n_semanas+int_semana+2 and semana<len(cts[id_ct].capacidade_iniciais):
        if duracao > ocupacao_max_semana * capacidade_med:
            duracao_alocar = ocupacao_max_semana * capacidade_med
        else:
            duracao_alocar = duracao

        if verificar_capacidades(id_of, semana, id_ov, capacidade_max, duracao_alocar,lista_verdes,lista_azuis):
            duracao = duracao - duracao_alocar

        if verificar_capacidades(id_of, semana, id_ov, capacidade_max, duracao_alocar,lista_verdes,lista_azuis)==False and semana==int_semana:
            return False

        semana += 1

    if duracao > 0:
        return False
    else:
        return True

def alocar_capacidade_rolante(id_of,int_semana,id_ov,ocupacao_max_semana,capacidade_max,lista_verdes,lista_azuis):

    duracao=ofs[id_of].duracao
    id_ct=ofs[id_of].id_ct
    capacidade_med=sum(cts[id_ct].capacidade_iniciais)/len(cts[id_ct].capacidade_iniciais)
    n_semanas=math.ceil(duracao / (ocupacao_max_semana * capacidade_med))

    if duracao==0:
        alocar(id_of, int_semana, id_ov, 0)

    while duracao>0 and int_semana<len(cts[id_ct].capacidade_iniciais):
        if duracao>ocupacao_max_semana*capacidade_med:
            duracao_alocar=ocupacao_max_semana*capacidade_med
        else:
            duracao_alocar=duracao

        if verificar_capacidades(id_of,int_semana,id_ov,capacidade_max,duracao_alocar,lista_verdes,lista_azuis):
            alocar(id_of,int_semana,id_ov,duracao_alocar)
            duracao=duracao-duracao_alocar

        int_semana+=1

def lista_ordenada(id_item):


    lista_ofs=items[id_item].id_ofs

    lista_ordenada=[]

    for index in range(len(lista_ofs)):
        id_of=lista_ofs[index]

        pos_alocar_max = len(lista_ordenada)
        pos_alocar_min = 0

        if any(x in ofs[id_of].precedencias for x in lista_ordenada) or any(x in ofs[id_of].sequencias for x in lista_ordenada):

            for pos_id in range(len(lista_ordenada)):
                id_verificar=lista_ordenada[pos_id]
                if id_of in ofs[id_verificar].precedencias:
                    if pos_alocar_max>=pos_id:
                        pos_alocar_max=pos_id
                if id_of in ofs[id_verificar].sequencias:
                    if pos_alocar_min>=pos_id:
                        pos_alocar_min=pos_id

            lista_ordenada.insert(pos_alocar_max,id_of)

        else:

            lista_ordenada.insert(0,id_of)

    return lista_ordenada

def print_wip_inicial(id_ovs):

    kpis = []

    wip_100 = 0
    count_ovs = 0

    for pos_ov in range(len(id_ovs)):

        id_ov = id_ovs[pos_ov]

        semana_max = 0
        n_ofs_semana_max = 0
        n_ofs_duracao = 0
        count_ofs = 0
        flag_duracao = 0
        flag_cliente = ""

        for pos_item in range(len(ovs[id_ov].id_items)):

            id_item = ovs[id_ov].id_items[pos_item]

            for pos_of in range(len(items[id_item].id_ofs)):

                id_of = items[id_item].id_ofs[pos_of]

                count_ofs += 1

                if len(ofs[id_of].id_alocada) > 0:

                    if max(ofs[id_of].id_alocada) > semana_max:
                        semana_max = max(ofs[id_of].id_alocada)

            for pos_of in range(len(items[id_item].id_ofs)):

                id_of = items[id_item].id_ofs[pos_of]

                if len(ofs[id_of].id_alocada) > 0:

                    if max(ofs[id_of].id_alocada) == semana_max:
                        n_ofs_semana_max += 1


            if n_ofs_duracao == 1 and count_ofs > 1:
                flag_duracao = 1

            flag_cliente = ovs[id_ov].sold_to

        if count_ofs == 0:

            minimizacao_wip = 0

        else:

            minimizacao_wip = n_ofs_semana_max / count_ofs

            count_ovs += 1

            if minimizacao_wip == 1:
                wip_100 += 1

            id_cliente = ovs[id_ov].id_cliente

            new_ov = {'ov': ovs[id_ov].cod_ov, 'semana de entrega': semana_max + semana_inicio_plano,
                      'minimizacao_wip': minimizacao_wip, 'alerta duracao < 60': flag_duracao,
                      'alerta cliente': flag_cliente, 'id_ov': id_ov,
                      'carga completa': clientes[id_cliente].carga_completa}

    if count_ovs!=0:

        new_kpi = {'kpi': wip_100 / count_ovs}
    else:
        new_kpi = {'kpi': 0}

    kpis.append(new_kpi)
    df_wip = pd.DataFrame(kpis)
    df_wip.to_csv('data/106. kpi wip AS IS.csv')


def print_data_entrega(id_ovs):

    ovs_output=[]
    centros_output=[]
    kpis=[]

    wip_100=0
    count_ovs=0


    for pos_ov in range(len(id_ovs)):

        id_ov=id_ovs[pos_ov]

        semana_max=0
        n_ofs_semana_max=0
        n_ofs_duracao=0
        count_ofs=0
        flag_duracao = 0
        flag_cliente = ""
        flag_ccs=""
        flag_cnm=""
        flag_crm=""

        count_ovs +=1

        for pos_item in range(len(ovs[id_ov].id_items)):

            id_item=ovs[id_ov].id_items[pos_item]

            if items[id_item].planeador=='CCS':
                flag_ccs='CCS'
            elif items[id_item].planeador=='CRM - Transformaçã':
                flag_crm = 'CRM'
            elif items[id_item].planeador=='CNM - Transformaçã':
                flag_cnm = 'CNM'

            for pos_of in range(len(items[id_item].id_ofs)):

                id_of = items[id_item].id_ofs[pos_of]

                count_ofs += 1

                if len(ofs[id_of].id_alocada) > 0:

                    if max(ofs[id_of].id_alocada) > semana_max:
                        semana_max = max(ofs[id_of].id_alocada)

            for pos_of in range(len(items[id_item].id_ofs)):

                id_of = items[id_item].id_ofs[pos_of]

                if len(ofs[id_of].id_alocada)>0:

                    if max(ofs[id_of].id_alocada) == semana_max:
                        n_ofs_semana_max+=1

                    if ofs[id_of].semana_min == semana_max and ofs[id_of].duracao<=60:
                        n_ofs_duracao+=1

            if n_ofs_duracao==1 and count_ofs>1:
                flag_duracao=1


            # if ovs[id_ov].id_cliente==11876:
            # flag_cliente=[ovs[id_ov].id_cliente]

            flag_cliente=ovs[id_ov].sold_to

            if count_ofs==0:

                minimizacao_wip=0

            else:

                minimizacao_wip=n_ofs_semana_max/count_ofs

                if minimizacao_wip==1:
                    wip_100+=1

            if flag_ccs!="":

                new_centro = {'id_ov': id_ov, 'area': flag_ccs}

                centros_output.append(new_centro)

            if flag_cnm!="":
                new_centro = {'id_ov': id_ov,'area': flag_cnm}

                centros_output.append(new_centro)

            if flag_crm!="":

                new_centro = {'id_ov': id_ov, 'area': flag_crm}

                centros_output.append(new_centro)

        id_cliente=ovs[id_ov].id_cliente

        if count_ofs!=0:

            new_ov = {'ov': ovs[id_ov].cod_ov, 'semana de entrega': semana_max + semana_inicio_plano,
                      'minimizacao_wip': minimizacao_wip, 'alerta duracao < 60': flag_duracao,
                      'alerta cliente': flag_cliente,
                      'key': str(ovs[id_ov].cod_ov).split('.')[0] + " - " + str(ovs[id_ov].data_desejada + semana_inicio_plano),
                      'carga completa': clientes[id_cliente].carga_completa, 'id_ov': id_ov}

            ovs_output.append(new_ov)

    try:
        temp= wip_100/count_ovs
    except:
        temp=0
    new_kpi={'kpi':temp}
    kpis.append(new_kpi)
    df_wip=pd.DataFrame(kpis)
    df_wip.to_csv('data/105. kpi wip.csv')

    df=pd.DataFrame(ovs_output)
    df.to_csv('data/103. semana entrega.csv')

    df_centros=pd.DataFrame(centros_output)
    df_centros.to_csv('data/107. centros ovs.csv')

    cumprimento_plano(ovs_output,correr_cumprimento)


def print_acabamentos():

    rows=[]

    for index in range(len(cts)):

        id_ct=cts[index].id

        if cts[id_ct].acabamento!='Total':

            for pos_semana in range(len(cts[id_ct].capacidade)):

                if cts[id_ct].capacidade[pos_semana]!=cts[id_ct].capacidade_iniciais[pos_semana]:

                    acabamento=1

                else:
                    acabamento=0

                new_row={'Centro de Trabalho' : cts[id_ct].nome,'semana':pos_semana+semana_inicio_plano,'acabamento':acabamento}

                rows.append(new_row)

    df=pd.DataFrame(rows)
    df.to_csv('data/104. acabamentos.csv')


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

def gerar_ficheiro_excel(planeador,id_ovs,wbk,wbkName,wks):

    for row in wks['A3:N' + str(len(wks['A']))]:
        for cell in row:
            cell.value = None

    id_prod = 0

    for index in range(len(id_ovs)):

        id_ov = id_ovs[index]

        lista_items = ovs[id_ov].id_items

        for pos_item in range(len(lista_items)):

            id_item = lista_items[pos_item]

            if items[id_item].planeador.split(' ')[0] == planeador:

                lista_ofs = lista_ordenada(id_item)

                items[id_item].id_ofs = lista_ofs

                for pos_of in range(len(lista_ofs)):

                    id_of = lista_ofs[pos_of]

                    if len(ofs[id_of].id_alocada)>0:

                        id_prod += 1

                        year = datetime.datetime.now().year
                        week_num = min(ofs[id_of].id_alocada) + semana_inicio_plano

                        date = datetime.date(year, 1, 1) + relativedelta(weeks=+week_num)
                        if len(str(date.month)) == 1:
                            month = "0" + str(date.month)
                        else:
                            month = date.month

                        if len(str(date.day)) == 1:
                            day = "0" + str(date.day)
                        else:
                            day = date.day

                        date_str = str(date.year) + str(month) + str(day)

                        wks.cell(row=id_prod + 2, column=1).value = id_prod
                        wks.cell(row=id_prod + 2, column=2).value = ovs[id_ov].cod_ov
                        wks.cell(row=id_prod + 2, column=3).value = items[id_item].cod_item
                        wks.cell(row=id_prod + 2, column=4).value = ofs[id_of].codigo_material
                        wks.cell(row=id_prod + 2, column=5).value = 3801
                        wks.cell(row=id_prod + 2, column=6).value = ofs[id_of].tipo_ordem
                        wks.cell(row=id_prod + 2, column=7).value = ofs[id_of].quantidade
                        wks.cell(row=id_prod + 2, column=8).value = "0001"
                        wks.cell(row=id_prod + 2, column=9).value = ""
                        wks.cell(row=id_prod + 2, column=10).value = date_str
                        wks.cell(row=id_prod + 2, column=11).value = ""
                        wks.cell(row=id_prod + 2, column=12).value = ""
                        wks.cell(row=id_prod + 2, column=13).value = "X"
                        wks.cell(row=id_prod + 2, column=14).value = "X"

    wbk.save(wbkName)
    wbk.close()

def gerar_output_sap(id_ovs):

    global semana_inicio_plano

    df_ofs = pd.read_csv('data/08.  ofs.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
    planeadores = df_ofs['Nome planejador'].dropna()
    planeadores = planeadores.unique()
    planeadores = planeadores.tolist()

    CCS=load_workbook("SAP/CCS_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx")
    CNM = load_workbook("SAP/CNM_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx")
    CRM = load_workbook("SAP/CRM_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx")

    wks_ccs = CCS.worksheets[0]
    wks_cnm = CNM.worksheets[0]
    wks_crm = CRM.worksheets[0]

    gerar_ficheiro_excel("CRM", id_ovs, CRM, "SAP/CRM_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx",wks_crm)
    gerar_ficheiro_excel("CNM", id_ovs, CNM, "SAP/CNM_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx",wks_cnm)
    gerar_ficheiro_excel("CCS", id_ovs, CCS, "SAP/CCS_ZPP_IMPORT_OF - Importação de OF's (PULL).xlsx",wks_ccs)

def gerar_output_sap_2(id_ovs):

    global semana_inicio_plano

    df_ofs = pd.read_csv('data/08.  ofs.csv', sep=",", encoding='iso-8859-1',error_bad_lines=False)
    planeadores = df_ofs['Nome planejador'].dropna()
    planeadores = planeadores.unique()
    planeadores = planeadores.tolist()

    for pos_planeador in range(len(planeadores)):

        planeador = planeadores[pos_planeador]
        planeador = planeador.split(' ')[0]
        output_final = []
        new_row = {"ID de Importação": "ID_PRROD",
                   "Ordem de Venda": "VBELN",
                   "Item da Ordem de Venda": "POSNR",
                   "Material": "MATNR",
                   "Centro": "WERKS",
                   "Tipo de Ordem": "AUART",
                   'Quantidade Total': "GAMNG",
                   'Versão de Produção': "VERID",
                   'Data de Inicio': "GSTRP",
                   "Data de fim": "GLTRP",
                   "Kanban": "ZZ_KANBAN",
                   "Texto Descritivo": "TEXTO",
                   "Consumos não Planeados": "ZZ_CONS",
                   "MTO": "MTO"}
        output_final.append(new_row)

        id_prod = 0

        for index in range(len(id_ovs)):

            id_ov = id_ovs[index]

            lista_items = ovs[id_ov].id_items

            for pos_item in range(len(lista_items)):

                id_item = lista_items[pos_item]

                if items[id_item].planeador == planeador:

                    lista_ofs = lista_ordenada(id_item)

                    items[id_item].id_ofs = lista_ofs

                    for pos_of in range(len(lista_ofs)):

                        id_of = lista_ofs[pos_of]

                        if len(ofs[id_of].id_alocada)>0:

                            id_prod += 1

                            year = datetime.datetime.now().year
                            week_num = min(ofs[id_of].id_alocada) + semana_inicio_plano

                            date = datetime.date(year, 1, 1) + relativedelta(weeks=+week_num)
                            if len(str(date.month)) == 1:
                                month = "0" + str(date.month)
                            else:
                                month = date.month

                            if len(str(date.day)) == 1:
                                day = "0" + str(date.day)
                            else:
                                day = date.day

                            date_str = str(date.year) + str(month) + str(day)

                            new_row = {"ID de Importação": id_prod,
                                       "Ordem de Venda": ovs[id_ov].cod_ov,
                                       "Item da Ordem de Venda": items[id_item].cod_item,
                                       "Material": ofs[id_of].codigo_material,
                                       "Centro": 3801,
                                       "Tipo de Ordem": ofs[id_of].tipo_ordem,
                                       'Quantidade Total': ofs[id_of].quantidade,
                                       'Versão de Produção': "0001",
                                       'Data de Inicio': "",
                                       "Data de fim": date_str,
                                       "Kanban": "",
                                       "Texto Descritivo": "",
                                       "Consumos não Planeados": "X",
                                       "MTO": "X"}

                            output_final.append(new_row)

        df_sap = pd.DataFrame(output_final)
        df_sap.to_excel('SAP/' + str(planeador) + '.xlsx', index=False, encoding='iso-8859-1')

def verificar_planeador(id_item):

    global lista_materials

    lista_ofs=items[id_item].id_ofs

    for id_of in lista_ofs:

        if 'EMBALAGEM' in ofs[id_of].descritivo_ct or int(ofs[id_of].codigo_material) in lista_materials:

            items[id_item].planeador=ofs[id_of].planeador

def atualizar_capacidade_ref_c(method):

    global semana_inicio_plano

    df_ocupacao=pd.read_csv('data/16. ocupacao ref c.csv',error_bad_lines=False)

    centro_trabalho=df_ocupacao['Centro trabalho'].tolist()

    df_ocupacao['Semana']=df_ocupacao['Semana']-semana_inicio_plano-1

    semanas=df_ocupacao['Semana'].tolist()

    ocupacoes=df_ocupacao['Capacidade'].tolist()

    for pos_ct in range(len(cts)):

        id_ct=cts[pos_ct].id

        if cts[id_ct].nome in centro_trabalho:

            for pos in range(len(semanas)):

                semana=semanas[pos]
                centro=centro_trabalho[pos]
                ocupacao=ocupacoes[pos]

                if semana>0 and centro==cts[id_ct].nome:

                    cts[id_ct].capacidade[semana]-=ocupacao*cts[id_ct].capacidade_iniciais[semana]

def importar_capacidade_cilindros():

    global pos_calandrados
    global pos_plyups
    global capacidade_calandrados
    global capacidade_plyups

    tamanho=len(cts[0].capacidade_iniciais)

    pos_calandrados=[]
    pos_plyups=[]
    capacidade_calandrados=[]
    capacidade_plyups=[]

    df_bom=pd.read_csv('data/19. BOM.csv',encoding='iso-8859-1',error_bad_lines=False)
    df_bom['Componente'] = pd.to_numeric(df_bom['Componente'], errors='coerce')
    df_bom['Material'] = pd.to_numeric(df_bom['Material'], errors='coerce')
    lista_material_of_calandrado=df_bom[df_bom['Componente'].isin(lista_calandrados)]
    lista_material_of_calandrado=lista_material_of_calandrado[['Material']]
    lista_material_of_calandrado=lista_material_of_calandrado['Material'].unique()
    lista_material_of_calandrado =lista_material_of_calandrado.tolist()

    lista_material_of_plyups = df_bom[df_bom['Componente'].isin(lista_plyups)]
    lista_material_of_plyups = lista_material_of_plyups[['Material']]
    lista_material_of_plyups=lista_material_of_plyups['Material'].unique()
    lista_material_of_plyups = lista_material_of_plyups.tolist()

    for index in range(len(lista_material_of_calandrado)):
        pos_calandrados.append(lista_material_of_calandrado[index])
        temp=[]
        for semana in range(tamanho):
            temp.append(4*26)
        capacidade_calandrados.insert(0,temp)

    for index in range(len(lista_material_of_plyups)):
        pos_plyups.append(lista_material_of_plyups[index])
        temp=[]
        for semana in range(tamanho):
            temp.append(3*26)
        capacidade_plyups.insert(0,temp)

def get_ids(ovs):

    id_ovs=[]

    for ov in ovs:
        id_ovs.append(ov.id)

    return id_ovs

def criar_capacidade_rolante():

    output=[]

    for ct in cts:

        acumulado=0

        for id_semana in range(len(ct.capacidade)):

            if ct.capacidade[id_semana]<0:

                acumulado=math.fabs(ct.capacidade[id_semana])

                ct.capacidade[id_semana] = 0

            else:

                remanescente=ct.capacidade[id_semana]

                if remanescente>=acumulado:
                    ct.capacidade[id_semana]-=acumulado
                else:
                    ct.capacidade[id_semana] -= remanescente
                    acumulado-=remanescente

            new_row={'centro de trabalho':ct.nome,'acabamento':ct.acabamento,'semana':id_semana+semana_inicio_plano,'carga ocupada':ct.capacidade[id_semana]}
            output.append(new_row)

    df=pd.DataFrame(output)
    df.to_csv('data/22. capacidade_rolante.csv',sep=',',encoding='ISO-8859-1')

def cumprimento_plano(ovs_dict,correr_cumprimento):


    df_cumprimento_plano=pd.read_csv('data/99. cumprimento.csv',error_bad_lines=False)
    ovs=df_cumprimento_plano['ov'].tolist()
    semana_algoritmo=df_cumprimento_plano['semana_algoritmo'].tolist()
    semana_planeada=df_cumprimento_plano['semana_plan_real'].tolist()

    df_ofs=pd.read_csv('data/08.  ofs.csv',sep=",",encoding='ISO-8859-1',error_bad_lines=False)

    df_ofs = df_ofs[df_ofs['Ordem Venda / Transferência'].notna()]
    df_ofs = df_ofs[df_ofs['Data-base do fim'].notna()]

    df_ofs['data'] = pd.to_datetime(df_ofs['Data-base do fim'], format='%d/%m/%Y',errors='coerce')
    df_ofs['semana']=df_ofs['data'].apply(lambda x: x.isocalendar()[1])

    ovs_global=df_ofs['Ordem Venda / Transferência'].unique().tolist()
    ovs_total=df_ofs['Ordem Venda / Transferência'].tolist()
    ofs_global=df_ofs['Ordem de produção / planeada'].tolist()
    semana_global=df_ofs['semana'].tolist()


    for linha in ovs_dict:
        if linha.get('ov') in ovs:
            posicao=ovs.index(linha.get('ov'))
            semana_algoritmo[posicao]=linha.get('semana de entrega')
        else:
            ovs.append(linha.get('ov'))
            semana_algoritmo.append(linha.get('semana de entrega'))
            semana_planeada.append(-1)

    if correr_cumprimento==1:

        print('A gerar cumprimento do plano')

        for ov in ovs_global:
            result=True
            semana_max=-1
            for of in ofs_global:
                posicao_of=ofs_global.index(of)
                if ovs_total[posicao_of]==ov:
                    if of<1600000000:
                        result=False
                        break
                    elif semana_global[posicao_of]>semana_max:
                        semana_max=semana_global[posicao_of]

            if result==True:
                if ov in ovs:
                    posicao_ov=ovs.index(ov)
                    semana_planeada[posicao_ov]=semana_max
                else:
                    ovs.append(ov)
                    semana_planeada.append(semana_max)
                    semana_algoritmo.append(-1)

    output=[]
    for posicao in range(len(ovs)):
        new_row={'ov':ovs[posicao],'semana_algoritmo':semana_algoritmo[posicao],'semana_plan_real':semana_planeada[posicao]}
        output.append(new_row)

    df=pd.DataFrame(output)
    df.to_csv('data/99. cumprimento.csv')


def alterar_centro_trabalho_amoltex():

    id_ct_anterior=-1
    for ct in cts:
        if ct.nome=='CRMLAMPL' and ct.acabamento=='Total':
            id_ct_anterior=ct.id
            break

    id_ct_atual=-1
    for ct in cts:
        if ct.nome=='CRMLAMPL' and ct.acabamento=='Printing + Manual':
            id_ct_atual=ct.id
            break

    for ov in ovs:
        if ov.sold_to=="AMOSEALTEX CORK COMPANY LIMITED":
            id_items=ov.id_items
            for id_item in id_items:
                id_ofs=items[id_item].id_ofs
                for id_of in id_ofs:
                    if ofs[id_of].id_ct==id_ct_anterior:
                        ofs[id_of].id_ct=id_ct_atual






